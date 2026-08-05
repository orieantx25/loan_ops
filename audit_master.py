# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from collections import Counter
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'Master sheet - Loans .xlsx', data_only=True, read_only=True)
ws = wb['Master data']

rows = []
headers = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    if row[0] is None and row[1] is None:
        continue
    rows.append(row)

print(f'Total data rows: {len(rows)}')
print(f'Total columns: {len([h for h in headers if h])}')
print()
print('=== ALL HEADERS ===')
for i, h in enumerate(headers):
    if h is not None:
        print(f'  {get_column_letter(i+1)} ({i}): {repr(h)[:90]}')

skip = {
    'Name', 'S. No', 'Mobile Number', 'Provisional ID',
    'SST Remarks', 'Pranjal Comments', 'SST Comments',
    'Co-lateral details', 'Column 37', 'Column 39', 'Column 40', 'Column 41',
    'Mentor Flag ', 'Intent to be Reverted to SST', 'Tentative Loan Date'
}

print()
print('=== UNIQUE VALUES FOR KEY COLUMNS ===')
for col_name in headers:
    if col_name is None or col_name in skip:
        continue
    idx = headers.index(col_name)
    vals = Counter()
    blank = 0
    for r in rows:
        v = r[idx] if idx < len(r) else None
        if v is None or (isinstance(v, str) and str(v).strip() == ''):
            blank += 1
        else:
            vals[str(v).strip()] += 1
    if len(vals) <= 100:
        print(f'\n--- {repr(col_name)[:70]} (blank={blank}, unique={len(vals)}) ---')
        for v, c in vals.most_common(50):
            print(f'  [{c:4d}] {v[:100]}')

# Vendor sharing analysis
vendor_cols = [
    'Shared to ICICI',
    'Shared to Propelld ',
    'Shared to Study4Buddy',
    'Shared to Poonawala Fincorp',
    'Shared to GyanDhan',
]
print('\n=== VENDOR MULTI-SHARE ANALYSIS ===')
vendor_idxs = []
for vc in vendor_cols:
    if vc in headers:
        vendor_idxs.append((vc, headers.index(vc)))
        print(f'Found: {repr(vc)}')

multi = Counter()
for r in rows:
    count = 0
    for name, idx in vendor_idxs:
        v = r[idx] if idx < len(r) else None
        if v is not None and str(v).strip().lower() in ('yes', 'y', 'true', '1'):
            count += 1
    multi[count] += 1

print('Vendor share count distribution:')
for k in sorted(multi):
    print(f'  {k} vendors: {multi[k]} students')

# Duplicate mobile / provisional ID
print('\n=== DUPLICATE KEY ANALYSIS ===')
mob_idx = headers.index('Mobile Number')
pid_idx = headers.index('Provisional ID')
mobs = Counter()
pids = Counter()
for r in rows:
    m = r[mob_idx]
    p = r[pid_idx]
    if m is not None:
        mobs[str(m).strip()] += 1
    if p is not None:
        pids[str(p).strip()] += 1
dup_mobs = {k: v for k, v in mobs.items() if v > 1}
dup_pids = {k: v for k, v in pids.items() if v > 1}
print(f'Unique mobiles: {len(mobs)}, duplicates: {len(dup_mobs)}')
print(f'Unique provisional IDs: {len(pids)}, duplicates: {len(dup_pids)}')
if dup_mobs:
    print('Sample duplicate mobiles:')
    for k, v in list(dup_mobs.items())[:10]:
        print(f'  {k}: {v}')
if dup_pids:
    print('Sample duplicate PIDs:')
    for k, v in list(dup_pids.items())[:10]:
        print(f'  {k}: {v}')

wb.close()
print('\nDONE')
