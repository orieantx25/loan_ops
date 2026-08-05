# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter

wb = openpyxl.load_workbook(r'Master sheet - Loans .xlsx', data_only=False, read_only=True)
ws = wb['Master data']

# Find formulas in Master data
formula_cols = Counter()
samples = {}
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=False)):
    if i == 0:
        headers = [c.value for c in row]
        continue
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith('='):
            col = get_column_letter(cell.column)
            formula_cols[col] += 1
            if col not in samples:
                samples[col] = (headers[cell.column-1] if cell.column-1 < len(headers) else '?', cell.value[:120])

print('=== FORMULAS IN MASTER DATA (first 50 rows) ===')
for col, cnt in formula_cols.most_common():
    h, f = samples[col]
    print(f'  {col} ({h}): count~{cnt} sample={f}')

wb.close()

# Also check merged cells and data validations briefly
wb2 = openpyxl.load_workbook(r'Master sheet - Loans .xlsx', data_only=False, read_only=False)
ws2 = wb2['Master data']
print(f'\nMerged cells in Master data: {len(ws2.merged_cells.ranges)}')
print(f'Data validations: {len(ws2.data_validations.dataValidation) if ws2.data_validations else 0}')

# Check conditional formatting
print(f'Conditional formatting rules: {len(ws2.conditional_formatting._cf_rules) if hasattr(ws2.conditional_formatting, "_cf_rules") else "n/a"}')

wb2.close()
