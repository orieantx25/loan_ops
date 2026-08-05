# -*- coding: utf-8 -*-
"""Extract Master Data → JSON for the web dashboard."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import json
from openpyxl import load_workbook
from pathlib import Path

SRC = Path(r'Master sheet - Loans .xlsx')
OUT = Path(r'web/src/data/students.json')

wb = load_workbook(SRC, data_only=True, read_only=True)
ws = wb['Master data']

headers = None
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = [str(h).replace('\n', ' ').strip() if h else f'col_{j}' for j, h in enumerate(row)]
        continue
    if row[1] is None or str(row[1]).strip() == '':
        continue
    rec = {}
    for j, h in enumerate(headers[:43]):
        v = row[j] if j < len(row) else None
        if isinstance(v, float) and v == int(v) and j != 7:  # keep scholarship float
            if j in (0, 2):  # sno, mobile
                v = int(v)
            elif j == 2:
                v = int(v)
        if hasattr(v, 'isoformat'):
            v = v.isoformat()
        rec[h] = v
    rows.append(rec)

wb.close()

# Normalize field names for cleaner TS access
FIELD_MAP = {
    'S. No': 'sno',
    'Name': 'name',
    'Mobile Number': 'mobile',
    'Provisional ID': 'provisionalId',
    'Campus': 'campus',
    'Need Loan (SST input)': 'needLoanSst',
    'Criticality  (SST INput)': 'criticality',
    'Scholarship': 'scholarship',
    'Loan Amt req': 'loanAmt',
    '12th percentage': 'pct12',
    'Co-Applicant Relation': 'coApplicant',
    'Occupation': 'occupation',
    'Annual Income': 'annualIncome',
    'Income Proof': 'incomeProof',
    'Any other loan': 'otherLoan',
    'Co-applcant Cibil': 'cibil',
    'Co-lateral': 'collateral',
    'Co-lateral details': 'collateralDetails',
    'SST Remarks': 'sstRemarks',
    'Pranjal Comments': 'pranjalComments',
    'Loan required (Latest)': 'loanRequired',
    'Shared to ICICI': 'sharedIcici',
    'Shared to Propelld': 'sharedPropelld',
    'Shared to Study4Buddy': 'sharedStudy4Buddy',
    'Shared to Poonawala Fincorp': 'sharedPoonawala',
    'Shared to GyanDhan': 'sharedGyandhan',
    'Initial Case Status': 'caseStatus',
    'CurrentCase Status (started on 31/7)': 'currentCaseStatus',
    'Tentative Loan Date': 'tentativeDate',
    'Intent to be Reverted to SST': 'intentRevert',
    'Mentor Flag': 'mentorFlag',
    'SST Comments': 'sstComments',
    'Loan Stage': 'loanStage',
    'Loan Status': 'loanStatus',
    'Need FLDG attention': 'needFldg',
    "Needs Vishwa's attention": 'needVishwa',
    'Special Vidyalaksmi attention': 'needVidyalakshmi',
    'Reason if not started': 'reasonNotStarted',
    'Banker Status': 'bankerStatus',
}

# Fix header keys that may have trailing spaces
def clean_key(k):
    return k.strip()

normalized = []
for r in rows:
    # Build by position-aware known headers from first row keys
    out = {}
    # Use positional extraction from original headers list
    keys = list(r.keys())
    def get_by_partial(*parts):
        for k in keys:
            kl = k.lower()
            if all(p.lower() in kl for p in parts):
                return r[k]
        return None

    def get_exactish(name):
        for k in keys:
            if clean_key(k).lower() == name.lower():
                return r[k]
        return None

    mobile = r.get('Mobile Number')
    if isinstance(mobile, float):
        mobile = int(mobile)

    shared_prop = None
    for k in keys:
        if 'propelld' in k.lower():
            shared_prop = r[k]
            break

    out = {
        'sno': r.get('S. No'),
        'name': r.get('Name'),
        'mobile': mobile,
        'provisionalId': r.get('Provisional ID'),
        'campus': r.get('Campus'),
        'needLoanSst': get_by_partial('Need Loan'),
        'criticality': get_by_partial('Criticality'),
        'scholarship': r.get('Scholarship'),
        'loanAmt': get_by_partial('Loan Amt'),
        'annualIncome': get_by_partial('Annual Income'),
        'incomeProof': get_by_partial('Income Proof'),
        'sstRemarks': get_by_partial('SST Remarks'),
        'pranjalComments': get_by_partial('Pranjal'),
        'loanRequired': get_by_partial('Loan required'),
        'sharedIcici': get_by_partial('Shared to ICICI') or get_exactish('Shared to ICICI'),
        'sharedPropelld': shared_prop,
        'sharedStudy4Buddy': get_by_partial('Study4Buddy'),
        'sharedPoonawala': get_by_partial('Poonawala'),
        'sharedGyandhan': get_by_partial('GyanDhan'),
        'caseStatus': get_by_partial('Initial Case'),
        'currentCaseStatus': get_by_partial('CurrentCase'),
        'tentativeDate': get_by_partial('Tentative'),
        'loanStage': get_by_partial('Loan Stage'),
        'loanStatus': get_by_partial('Loan Status'),
        'needFldg': get_by_partial('FLDG'),
        'needVishwa': get_by_partial('Vishwa'),
        'needVidyalakshmi': get_by_partial('Vidyalaksmi') or get_by_partial('Vidyalakshmi'),
        'reasonNotStarted': get_by_partial('Reason if not'),
    }
    normalized.append(out)

OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(normalized, f, ensure_ascii=False, indent=None, default=str)

print(f'Wrote {len(normalized)} students → {OUT}')
# sanity
yes = sum(1 for s in normalized if str(s.get('loanRequired') or '').strip().lower() == 'yes')
print(f'loanRequired=Yes: {yes}')
print('sample keys', list(normalized[0].keys()))
print('sample', {k: normalized[0][k] for k in list(normalized[0])[:8]})
