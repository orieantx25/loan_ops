# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'Master sheet - Loans .xlsx', data_only=False, read_only=False)

sheets = [
    'Data ', 'Data Summary', 'Dropdowns', 'Pivot Table 2',
    'Need FLDG attention', 'Student deferring loan requirem',
    'Need Vidyalakshmi Attention', 'Specific need students',
    'Students who want loan from nex', 'SummaryView', 'Sheet13', 'Phase 2'
]

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        print(f'\nSKIP missing: {sheet_name}')
        continue
    ws = wb[sheet_name]
    print(f'\n========== [{sheet_name}] max_row={ws.max_row} max_col={ws.max_column} ==========')
    max_r = min(50, ws.max_row or 50)
    max_c = min(22, ws.max_column or 22)
    for row_idx in range(1, max_r + 1):
        cells = []
        for col_idx in range(1, max_c + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                v = str(cell.value).replace('\n', ' ')[:140]
                cells.append(f'{get_column_letter(col_idx)}={v}')
        if cells:
            print(f'R{row_idx}: {" | ".join(cells)}')

# Tables defined
print('\n========== DEFINED TABLES ==========')
for name in wb.sheetnames:
    ws = wb[name]
    if hasattr(ws, 'tables') and ws.tables:
        for tname, table in ws.tables.items():
            print(f'  Sheet[{name}] Table[{tname}] ref={table.ref}')

# Named styles / defined names
print('\n========== DEFINED NAMES ==========')
for dn in wb.defined_names.values():
    try:
        print(f'  {dn.name}: {dn.attr_text}')
    except Exception as e:
        print(f'  {dn.name}: ERR {e}')

wb.close()
print('\nDONE')
