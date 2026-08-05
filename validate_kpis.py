# -*- coding: utf-8 -*-
"""Validate KPI parity by simulating Helper logic in Python."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import Counter

SRC = r'Master sheet - Loans .xlsx'
wb = load_workbook(SRC, data_only=True, read_only=True)
ws = wb['Master data']
headers = None
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    if row[1] is None:
        continue
    rows.append(row)
wb.close()

idx = {h: i for i, h in enumerate(headers) if h}

def g(r, h):
    i = idx.get(h)
    return r[i] if i is not None and i < len(r) else None

def is_yes(v):
    return v is not None and str(v).strip().lower() == 'yes'

def norm_campus(c):
    if c is None:
        return ''
    t = str(c).strip()
    m = {'SSHAE': 'SSAHE', 'ssHAE': 'SSAHE', 'ssahe': 'SSAHE', 'SSAHE': 'SSAHE', 'ADYPU': 'ADYPU'}
    return m.get(t, t.upper())

camp_map = {'SSHAE': 'SSAHE', 'ssHAE': 'SSAHE', 'ssahe': 'SSAHE'}

def vendor_count(r):
    cols = ['Shared to ICICI', 'Shared to Propelld ', 'Shared to Study4Buddy',
            'Shared to Poonawala Fincorp', 'Shared to GyanDhan']
    return sum(1 for c in cols if is_yes(g(r, c)))

def canonical(r):
    H = str(g(r, 'Loan required\n(Latest)') or '').strip()
    U = str(g(r, 'Initial Case Status') or '').strip()
    V = str(g(r, 'Loan Stage ') or '').strip()
    W = str(g(r, 'Loan Status') or '').strip()
    Hl, Ul, Vl, Wl = H.lower(), U.lower(), V.lower(), W.lower()
    Q = vendor_count(r)
    if Hl == 'refund' or Ul == 'refund case':
        return 'Refund'
    if Hl == 'dnp' or Ul == 'dnp':
        return 'DNP'
    if Vl == 'disbursed' or 'disbursed' in W:
        return 'Disbursed'
    if Vl == 'loan proccessed/accepted' or 'sanctioned' in Wl:
        return 'Sanctioned'
    if Vl == 'rejected' or 'not eligbile' in W or 'not eligible' in Wl:
        return 'Rejected'
    if 'Docs Pending' in W:
        return 'Documents Pending'
    if Hl == 'yes' and Vl == 'ongoing':
        return 'Processing'
    if Hl == 'yes' and Q > 0 and (V == '' or Vl == 'not even started'):
        return 'Vendor Assigned'
    if Hl == 'yes' and Vl == 'not even started':
        return 'Need Loan'
    if Hl == 'yes':
        return 'Loan Started'
    if Hl == 'not sure':
        return 'Interested'
    if Hl == 'no' or Vl == 'not required':
        return 'Not Required'
    return 'Unclassified'

stages = Counter()
pipeline = 0
fldg = vidya = crit = dup = 0
shared = Counter()
overlap = Counter()
vendors = ['Shared to ICICI', 'Shared to Propelld ', 'Shared to Study4Buddy',
           'Shared to Poonawala Fincorp', 'Shared to GyanDhan']
vnames = ['ICICI', 'Propelld', 'Study4Buddy', 'Poonawala', 'GyanDhan']

for r in rows:
    st = canonical(r)
    stages[st] += 1
    if is_yes(g(r, 'Loan required\n(Latest)')) or str(g(r, 'Loan required\n(Latest)') or '').strip().lower() == 'yes':
        # already is_yes
        pass
    lr = g(r, 'Loan required\n(Latest)')
    if is_yes(lr):
        pipeline += 1
    if is_yes(g(r, 'Need FLDG attention')):
        fldg += 1
    if is_yes(g(r, 'Special Vidyalaksmi attention')):
        vidya += 1
    critity = str(g(r, 'Criticality \n(SST INput)') or '').strip().lower()
    if critity == 'critical' or is_yes(g(r, 'Need FLDG attention')) or is_yes(g(r, 'Special Vidyalaksmi attention')) or str(g(r, 'Initial Case Status') or '').strip().lower() == 'risk':
        crit += 1
    vc = vendor_count(r)
    if vc >= 2:
        dup += 1
    flags = [is_yes(g(r, c)) for c in vendors]
    for i, f in enumerate(flags):
        if f:
            shared[vnames[i]] += 1
    for i in range(5):
        for j in range(i+1, 5):
            if flags[i] and flags[j]:
                overlap[(vnames[i], vnames[j])] += 1

print('=== VALIDATION KPIs ===')
print(f'Total students: {len(rows)}')
print(f'Need Loan (Latest=Yes): {pipeline}')
print(f'FLDG: {fldg}')
print(f'Vidyalakshmi: {vidya}')
print(f'Critical Flag (composite): {crit}')
print(f'Duplicate vendors: {dup}')
print(f'Shared: {dict(shared)}')
print(f'Avg vendors/shared: {sum(shared.values())/sum(1 for r in rows if vendor_count(r)>=1):.2f}')
print()
print('Canonical stages:')
for k, v in stages.most_common():
    print(f'  {k}: {v}')
print()
print('Overlap sample:')
for k, v in sorted(overlap.items(), key=lambda x: -x[1])[:10]:
    print(f'  {k}: {v}')

# Campus
camp = Counter(norm_campus(g(r, 'Campus')) for r in rows if g(r, 'Campus'))
print('Campus:', dict(camp))

# Pipeline breakdown
print()
print('Pipeline by stage:')
pstage = Counter()
for r in rows:
    if is_yes(g(r, 'Loan required\n(Latest)')):
        pstage[canonical(r)] += 1
for k, v in pstage.most_common():
    print(f'  {k}: {v}')
