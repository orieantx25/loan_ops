# -*- coding: utf-8 -*-
"""
uGSOT Loan Operations Management System — Workbook Builder
Formula-driven architecture for Google Sheets (also opens in Excel 365).
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from collections import OrderedDict
from datetime import datetime

SRC = r'Master sheet - Loans .xlsx'
OUT = r'uGSOT_Loan_Operations_System.xlsx'

# ── Brand tokens ──────────────────────────────────────────────
RED = 'E31C24'
BLACK = '111111'
DARK = '222222'
BG = 'F6F6F6'
WHITE = 'FFFFFF'
BORDER = 'E5E5E5'
MUTED = '6B6B6B'
LIGHT_RED = 'FDECEC'
CARD_SHADOW = 'F0F0F0'
GREEN = '1B7A4E'
AMBER = 'B45309'
GREY_HEADER = '2B2B2B'

thin = Border(
    left=Side(style='thin', color=BORDER),
    right=Side(style='thin', color=BORDER),
    top=Side(style='thin', color=BORDER),
    bottom=Side(style='thin', color=BORDER),
)
thin_dark = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

fill_red = PatternFill('solid', fgColor=RED)
fill_black = PatternFill('solid', fgColor=BLACK)
fill_dark = PatternFill('solid', fgColor=DARK)
fill_bg = PatternFill('solid', fgColor=BG)
fill_white = PatternFill('solid', fgColor=WHITE)
fill_header = PatternFill('solid', fgColor=GREY_HEADER)
fill_light_red = PatternFill('solid', fgColor=LIGHT_RED)
fill_card = PatternFill('solid', fgColor=WHITE)
fill_kpi_label = PatternFill('solid', fgColor='FAFAFA')

font_brand = Font(name='Calibri', size=22, bold=True, color=RED)
font_title = Font(name='Calibri', size=16, bold=True, color=BLACK)
font_h = Font(name='Calibri', size=11, bold=True, color=WHITE)
font_label = Font(name='Calibri', size=9, color=MUTED)
font_kpi = Font(name='Calibri', size=20, bold=True, color=BLACK)
font_kpi_red = Font(name='Calibri', size=20, bold=True, color=RED)
font_body = Font(name='Calibri', size=10, color=BLACK)
font_small = Font(name='Calibri', size=9, color=DARK)
font_section = Font(name='Calibri', size=12, bold=True, color=BLACK)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill_header
        cell.font = font_h
        cell.alignment = center
        cell.border = thin


def autosize(ws, min_w=8, max_w=28, cols=None):
    if cols is None:
        cols = range(1, (ws.max_column or 1) + 1)
    for col in cols:
        letter = get_column_letter(col)
        maxlen = 8
        for row in ws.iter_rows(min_col=col, max_col=col, max_row=min(60, ws.max_row or 1)):
            for cell in row:
                if cell.value is not None:
                    maxlen = max(maxlen, min(max_w, len(str(cell.value).split('\n')[0])))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, maxlen + 2))


def paint_bg(ws, rows=80, cols=20):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = fill_bg


# ── Load source Master ────────────────────────────────────────
print('Loading source workbook...')
src = load_workbook(SRC, data_only=True, read_only=True)
src_ws = src['Master data']

headers = []
master_rows = []
for i, row in enumerate(src_ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
        continue
    if row[1] is None or str(row[1]).strip() == '':
        continue
    master_rows.append(list(row[:43]))  # A-AQ

src.close()
print(f'  Master rows: {len(master_rows)}, cols: {len([h for h in headers if h])}')

# Normalize header display names (flatten newlines)
HDR = []
for h in headers[:43]:
    if h is None:
        HDR.append('')
    else:
        HDR.append(str(h).replace('\n', ' ').strip())

N = len(master_rows)
BUFFER = max(500, N + 200)  # formula fill buffer
LAST = N + 1  # last data row on sheets (row 1 = header)

print(f'  Formula buffer rows: {BUFFER}')

# ══════════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────────
# SHEET: 01_Master_Data
# ──────────────────────────────────────────────────────────────
print('Building 01_Master_Data...')
ws = wb.active
ws.title = '01_Master_Data'
ws.sheet_properties.tabColor = '666666'

for c, h in enumerate(HDR, 1):
    cell = ws.cell(1, c, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.alignment = center
    cell.border = thin

for r_i, row in enumerate(master_rows, 2):
    for c, val in enumerate(row, 1):
        # Strip formulas — values only. For S.No use sequential.
        if c == 1:
            ws.cell(r_i, c, r_i - 1)
        else:
            # CurrentCase Status (col 28 / AB): keep value as-is from data_only
            v = val
            if isinstance(v, float) and v == int(v) and c in (3,):  # mobile as int text later
                pass
            cell = ws.cell(r_i, c, v)
            cell.font = font_body
            cell.border = thin

# Mobile as text-friendly number format
for r in range(2, N + 2):
    cell = ws.cell(r, 3)
    if cell.value is not None:
        try:
            cell.number_format = '0'
        except Exception:
            pass

ws.freeze_panes = 'B2'
ws.auto_filter.ref = f'A1:AQ{N+1}'
autosize(ws, max_w=22)
ws.column_dimensions['S'].width = 30
ws.column_dimensions['T'].width = 30

# Named logical columns (1-indexed) for formula refs:
# A=1 SNo B=2 Name C=3 Mobile D=4 ProvID E=5 Campus F=6 NeedLoanSST G=7 Criticality
# H=8 Scholarship I=9 LoanAmt J=10 Pct12 K=11 CoApp L=12 Occ M=13 Income N=14 IncProof
# O=15 OtherLoan P=16 Cibil Q=17 Collateral R=18 CollDet S=19 SSTRem T=20 Pranjal
# U=21 LoanReqLatest V=22 ICICI W=23 Propelld X=24 Study4 Y=25 Poonawala Z=26 GyanDhan
# AA=27 InitCase AB=28 CurrCase AC=29 TentDate AD=30 Intent AE=31 Mentor AF=32 SSTComm
# AG=33 LoanStage AH=34 LoanStatus AI=35 FLDG AJ=36 Vishwa AK=37 Vidya AL=38 Reason
# AM=39 Col37 AN=40 Banker AO=41 Col39 AP=42 Col40 AQ=43 Col41

MD = "'01_Master_Data'"

# ──────────────────────────────────────────────────────────────
# SHEET: 02_Configuration
# ──────────────────────────────────────────────────────────────
print('Building 02_Configuration...')
cfg = wb.create_sheet('02_Configuration')
cfg.sheet_properties.tabColor = RED

cfg['A1'] = 'uGSOT LOAN OPERATIONS — CONFIGURATION'
cfg['A1'].font = font_brand
cfg.merge_cells('A1:F1')
cfg['A2'] = 'Edit this sheet only to add vendors, stages, mappings. Do not edit formulas on other sheets.'
cfg['A2'].font = font_label
cfg.merge_cells('A2:F2')

# --- Vendors ---
cfg['A4'] = 'VENDORS'
cfg['A4'].font = font_section
cfg['A4'].fill = fill_light_red
vend_headers = ['VendorID', 'VendorName', 'MasterCol', 'MasterHeader', 'Active', 'SortOrder', 'StatusToken']
for i, h in enumerate(vend_headers, 1):
    cell = cfg.cell(5, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin

vendors = [
    (1, 'ICICI', 'V', 'Shared to ICICI', 'Yes', 1, 'Processing by Icici|Loan sanctioned by Icici'),
    (2, 'Propelld', 'W', 'Shared to Propelld', 'Yes', 2, 'Processing by Propelld|Loan sanctioned by Propelld|Loan Sanctioned by Proplld'),
    (3, 'Study4Buddy', 'X', 'Shared to Study4Buddy', 'Yes', 3, 'Processing by Study4Buddy'),
    (4, 'Poonawala Fincorp', 'Y', 'Shared to Poonawala Fincorp', 'Yes', 4, 'Processing by Poonawala Fincorp'),
    (5, 'GyanDhan', 'Z', 'Shared to GyanDhan', 'Yes', 5, 'Processing by GyanDhan|Loan sanctioned by gyan dhan'),
    (6, 'PM Vidyalakshmi', '', 'Special Vidyalaksmi / Loan Status', 'Yes', 6, 'Processing by PM vidyalakshmi'),
    (7, 'Other Bank', '', 'Loan Status token', 'Yes', 7, 'Processing at other bank|Loan sanctioned by other bank'),
    (8, 'State Scheme', '', 'Loan Status token', 'Yes', 8, 'State Scheme'),
    (9, 'Auxilo', '', '', 'No', 9, ''),
    (10, 'InCred', '', '', 'No', 10, ''),
]
for r, row in enumerate(vendors, 6):
    for c, v in enumerate(row, 1):
        cell = cfg.cell(r, c, v)
        cell.border = thin
        cell.font = font_body

# --- Stages ---
cfg['A18'] = 'CANONICAL STAGES (Funnel Order)'
cfg['A18'].font = font_section
cfg['A18'].fill = fill_light_red
stage_h = ['StageID', 'StageName', 'FunnelOrder', 'FunnelCore', 'DashboardKPI', 'Color']
for i, h in enumerate(stage_h, 1):
    cell = cfg.cell(19, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin

stages = [
    (1, 'Need Loan', 1, 'Yes', 'Yes', RED),
    (2, 'Interested', 2, 'No', 'No', '666666'),
    (3, 'Loan Started', 3, 'Yes', 'Yes', '333333'),
    (4, 'Documents Pending', 4, 'Yes', 'Yes', AMBER),
    (5, 'Vendor Assigned', 5, 'No', 'No', '666666'),
    (6, 'Processing', 6, 'Yes', 'Yes', '1D4ED8'),
    (7, 'Approved', 7, 'Yes', 'Yes', GREEN),
    (8, 'Sanctioned', 8, 'Yes', 'No', GREEN),
    (9, 'Disbursed', 9, 'Yes', 'Yes', GREEN),
    (10, 'Completed', 10, 'No', 'No', '111111'),
    (11, 'Rejected', 90, 'No', 'Yes', RED),
    (12, 'Refund', 91, 'No', 'No', MUTED),
    (13, 'DNP', 92, 'No', 'No', MUTED),
    (14, 'Not Required', 93, 'No', 'No', MUTED),
    (15, 'Unclassified', 99, 'No', 'No', MUTED),
]
for r, row in enumerate(stages, 20):
    for c, v in enumerate(row, 1):
        cell = cfg.cell(r, c, v)
        cell.border = thin
        cell.font = font_body

# --- Campus map ---
cfg['A38'] = 'CAMPUS NORMALIZATION'
cfg['A38'].font = font_section
cfg['A38'].fill = fill_light_red
for i, h in enumerate(['RawValue', 'Normalized', 'Active'], 1):
    cell = cfg.cell(39, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin
campuses = [
    ('SSAHE', 'SSAHE', 'Yes'),
    ('SSHAE', 'SSAHE', 'Yes'),
    ('ssHAE', 'SSAHE', 'Yes'),
    ('ssahe', 'SSAHE', 'Yes'),
    ('ADYPU', 'ADYPU', 'Yes'),
]
for r, row in enumerate(campuses, 40):
    for c, v in enumerate(row, 1):
        cell = cfg.cell(r, c, v)
        cell.border = thin

# --- Reason buckets ---
cfg['A48'] = 'REASON BUCKETS (keyword → bucket)'
cfg['A48'].font = font_section
cfg['A48'].fill = fill_light_red
for i, h in enumerate(['Keyword', 'Bucket'], 1):
    cell = cfg.cell(49, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin
reasons = [
    ('DNP', 'Unreachable / DNP'),
    ('unreachable', 'Unreachable / DNP'),
    ('interest', 'High Interest / Terms'),
    ('eligib', 'Eligibility / Income'),
    ('income', 'Eligibility / Income'),
    ('Low', 'Eligibility / Income'),
    ('start', 'Process Yet to Start'),
    ('Newer', 'Process Yet to Start'),
    ('Fresh', 'Process Yet to Start'),
    ('Not sure', 'Undecided'),
    ('reject', 'Rejected Elsewhere'),
    ('SBI', 'Rejected Elsewhere'),
    ('misunderstand', 'Miscommunication'),
    ('automatic', 'Miscommunication'),
    ('waiting', 'Family / Dependency'),
    ('sister', 'Family / Dependency'),
    ('father', 'Family / Dependency'),
    ('call', 'Bank Communication'),
    ('PAN', 'Documentation'),
    ('doc', 'Documentation'),
    ('info', 'Insufficient Info'),
]
for r, row in enumerate(reasons, 50):
    for c, v in enumerate(row, 1):
        cell = cfg.cell(r, c, v)
        cell.border = thin

# --- Ageing ---
cfg['D38'] = 'AGEING BUCKETS'
cfg['D38'].font = font_section
cfg['D38'].fill = fill_light_red
for i, h in enumerate(['Bucket', 'MinDays', 'MaxDays'], 4):
    cell = cfg.cell(39, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin
ageing = [
    ('0–3 Days', 0, 3),
    ('4–7 Days', 4, 7),
    ('8–15 Days', 8, 15),
    ('16–30 Days', 16, 30),
    ('30+ Days', 31, 9999),
    ('Unknown', -1, -1),
]
for r, row in enumerate(ageing, 40):
    for c, v in enumerate(row, 4):
        cell = cfg.cell(r, c, v)
        cell.border = thin

# --- Risk categories ---
cfg['D48'] = 'RISK FLAGS'
cfg['D48'].font = font_section
for i, h in enumerate(['FlagName', 'MasterCol', 'YesValue'], 4):
    cell = cfg.cell(49, i, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin
risks = [
    ('Need FLDG', 'AI', 'Yes'),
    ('Need Vishwa Attention', 'AJ', 'Yes'),
    ('Need Vidyalakshmi', 'AK', 'Yes'),
    ('Critical', 'G', 'Critical'),
    ('Refund', 'U', 'Refund'),
    ('Risk Case', 'AA', 'Risk'),
]
for r, row in enumerate(risks, 50):
    for c, v in enumerate(row, 4):
        cell = cfg.cell(r, c, v)
        cell.border = thin

# Brand colors reference
cfg['A74'] = 'BRAND COLORS'
cfg['A74'].font = font_section
cfg['A75'] = 'Primary Red'
cfg['B75'] = '#E31C24'
cfg['B75'].fill = fill_red
cfg['B75'].font = Font(color='FFFFFF', bold=True)
cfg['A76'] = 'Black'
cfg['B76'] = '#111111'
cfg['A77'] = 'Background'
cfg['B77'] = '#F6F6F6'
cfg['A78'] = 'Border'
cfg['B78'] = '#E5E5E5'

# Filter defaults
cfg['A80'] = 'DASHBOARD FILTER DEFAULTS'
cfg['A80'].font = font_section
cfg['A81'] = 'FilterAllToken'
cfg['B81'] = '(All)'
cfg['A82'] = 'AsOfDate'
cfg['B82'] = datetime(2026, 8, 3).date()
cfg['A83'] = 'AdmissionCycle'
cfg['B83'] = '2026 Cycle'

autosize(cfg, max_w=36)

# ──────────────────────────────────────────────────────────────
# SHEET: 03_Helper  (row-aligned processing engine)
# ──────────────────────────────────────────────────────────────
print('Building 03_Helper...')
hp = wb.create_sheet('03_Helper')
hp.sheet_properties.tabColor = '1D4ED8'

helper_cols = [
    'RowNum',              # A
    'StudentKey',          # B
    'StudentName',         # C
    'Mobile',              # D
    'ProvisionalID',       # E
    'CampusRaw',           # F
    'Campus',              # G
    'LoanRequired',        # H
    'NeedLoanSST',         # I
    'Criticality',         # J
    'Scholarship',         # K
    'SharedICICI',         # L
    'SharedPropelld',      # M
    'SharedStudy4Buddy',   # N
    'SharedPoonawala',     # O
    'SharedGyanDhan',      # P
    'VendorCount',         # Q
    'DuplicateVendorFlag', # R
    'VendorsApplied',      # S
    'PrimaryVendor',       # T
    'CaseStatus',          # U
    'LoanStageRaw',        # V
    'LoanStatusRaw',       # W
    'CanonicalStage',      # X
    'ApprovalStatus',      # Y
    'DisbursementStatus',  # Z
    'RiskCategory',        # AA
    'CriticalFlag',        # AB
    'NeedFLDG',            # AC
    'NeedVidyalakshmi',    # AD
    'NeedVishwa',          # AE
    'ReasonRaw',           # AF
    'ReasonBucket',        # AG
    'TentativeDate',       # AH
    'PendingDays',         # AI
    'AgeingBucket',        # AJ
    'IsLoanPipeline',      # AK  Yes if Latest=Yes
    'PranjalComments',     # AL
    'SSTRemarks',          # AM
    'LoanAmt',             # AN
    'Income',              # AO
    'ProcessingByICICI',   # AP  from Loan Status parse
    'ProcessingByPropelld',# AQ
    'ProcessingByGyanDhan',# AR
    'ProcessingByVidya',   # AS
    'ProcessingByOther',   # AT
    'IsSanctioned',        # AU
    'IsDocsPending',       # AV
    'IsRejectedStatus',    # AW
]

for c, h in enumerate(helper_cols, 1):
    cell = hp.cell(1, c, h)
    cell.fill = fill_header
    cell.font = font_h
    cell.alignment = center
    cell.border = thin

# Fill formulas for each data row (+ small buffer empty check)
# Using Excel/Sheets compatible formulas

def yes_norm(col_letter, r):
    """Normalize Yes check against Master."""
    return f'IF(LOWER(TRIM({MD}!{col_letter}{r}&""))="yes",1,0)'

for r in range(2, N + 2):
    # A RowNum
    hp.cell(r, 1, f'=IF({MD}!B{r}="","",{r-1})')
    # B StudentKey — Mobile primary, Provisional ID fallback
    hp.cell(r, 2,
        f'=IF({MD}!B{r}="","",'
        f'IF({MD}!C{r}<>"","MOB|"&TEXT({MD}!C{r},"0"),'
        f'IF({MD}!D{r}<>"","PID|"&TRIM({MD}!D{r}&""),"")))')
    # C Name
    hp.cell(r, 3, f'=IF({MD}!B{r}="","",{MD}!B{r})')
    # D Mobile
    hp.cell(r, 4, f'=IF({MD}!B{r}="","",{MD}!C{r})')
    # E ProvID
    hp.cell(r, 5, f'=IF({MD}!B{r}="","",{MD}!D{r})')
    # F CampusRaw
    hp.cell(r, 6, f'=IF({MD}!B{r}="","",{MD}!E{r})')
    # G Campus normalized
    hp.cell(r, 7,
        f'=IF({MD}!B{r}="","",'
        f'IFERROR(VLOOKUP(TRIM({MD}!E{r}&""),\'02_Configuration\'!$A$40:$B$44,2,FALSE),'
        f'UPPER(TRIM({MD}!E{r}&""))))')
    # H LoanRequired Latest (normalized)
    hp.cell(r, 8,
        f'=IF({MD}!B{r}="","",'
        f'IF({MD}!U{r}="","",TRIM({MD}!U{r}&"")))')
    # I NeedLoan SST
    hp.cell(r, 9, f'=IF({MD}!B{r}="","",TRIM({MD}!F{r}&""))')
    # J Criticality
    hp.cell(r, 10, f'=IF({MD}!B{r}="","",TRIM({MD}!G{r}&""))')
    # K Scholarship
    hp.cell(r, 11, f'=IF({MD}!B{r}="","",{MD}!H{r})')
    # L-P Shared flags normalized 0/1
    hp.cell(r, 12, f'=IF({MD}!B{r}="","",{yes_norm("V", r)})')
    hp.cell(r, 13, f'=IF({MD}!B{r}="","",{yes_norm("W", r)})')
    hp.cell(r, 14, f'=IF({MD}!B{r}="","",{yes_norm("X", r)})')
    hp.cell(r, 15, f'=IF({MD}!B{r}="","",{yes_norm("Y", r)})')
    hp.cell(r, 16, f'=IF({MD}!B{r}="","",{yes_norm("Z", r)})')
    # Q VendorCount
    hp.cell(r, 17, f'=IF(A{r}="","",L{r}+M{r}+N{r}+O{r}+P{r})')
    # R DuplicateVendorFlag
    hp.cell(r, 18, f'=IF(A{r}="","",IF(Q{r}>1,"Yes","No"))')
    # S VendorsApplied TEXTJOIN
    hp.cell(r, 19,
        f'=IF(A{r}="","",TEXTJOIN(", ",TRUE,'
        f'IF(L{r}=1,"ICICI",""),IF(M{r}=1,"Propelld",""),IF(N{r}=1,"Study4Buddy",""),'
        f'IF(O{r}=1,"Poonawala Fincorp",""),IF(P{r}=1,"GyanDhan","")))')
    # T PrimaryVendor — first shared, else from status
    hp.cell(r, 20,
        f'=IF(A{r}="","",'
        f'IF(L{r}=1,"ICICI",IF(M{r}=1,"Propelld",IF(P{r}=1,"GyanDhan",'
        f'IF(N{r}=1,"Study4Buddy",IF(O{r}=1,"Poonawala Fincorp",'
        f'IF(ISNUMBER(SEARCH("GyanDhan",W{r}&"")),"GyanDhan",'
        f'IF(ISNUMBER(SEARCH("Icici",W{r}&"")),"ICICI",'
        f'IF(ISNUMBER(SEARCH("Propelld",W{r}&"")),"Propelld",'
        f'IF(ISNUMBER(SEARCH("vidyalakshmi",LOWER(W{r}&"")),"PM Vidyalakshmi","—"))))))))))')
    # U CaseStatus
    hp.cell(r, 21, f'=IF(A{r}="","",TRIM({MD}!AA{r}&""))')
    # V LoanStageRaw
    hp.cell(r, 22, f'=IF(A{r}="","",TRIM({MD}!AG{r}&""))')
    # W LoanStatusRaw
    hp.cell(r, 23, f'=IF(A{r}="","",TRIM({MD}!AH{r}&""))')

    # X CanonicalStage — priority mapping
    hp.cell(r, 24,
        f'=IF(A{r}="","",'
        f'IF(OR(LOWER(H{r})="refund",LOWER(U{r})="refund case"),"Refund",'
        f'IF(OR(LOWER(H{r})="dnp",LOWER(U{r})="dnp"),"DNP",'
        f'IF(OR(LOWER(V{r})="disbursed",ISNUMBER(SEARCH("Disbursed",W{r}&""))),"Disbursed",'
        f'IF(OR(LOWER(V{r})="loan proccessed/accepted",ISNUMBER(SEARCH("sanctioned",LOWER(W{r}&"")))),"Sanctioned",'
        f'IF(OR(LOWER(V{r})="rejected",ISNUMBER(SEARCH("Not eligbile",W{r}&"")),ISNUMBER(SEARCH("Not eligible",W{r}&""))),"Rejected",'
        f'IF(ISNUMBER(SEARCH("Docs Pending",W{r}&"")),"Documents Pending",'
        f'IF(AND(LOWER(H{r})="yes",LOWER(V{r})="ongoing"),"Processing",'
        f'IF(AND(LOWER(H{r})="yes",Q{r}>0,OR(V{r}="",LOWER(V{r})="not even started")),"Vendor Assigned",'
        f'IF(AND(LOWER(H{r})="yes",LOWER(V{r})="not even started"),"Need Loan",'
        f'IF(LOWER(H{r})="yes","Loan Started",'
        f'IF(LOWER(H{r})="not sure","Interested",'
        f'IF(OR(LOWER(H{r})="no",LOWER(V{r})="not required"),"Not Required",'
        f'"Unclassified")))))))))))))')

    # Y ApprovalStatus
    hp.cell(r, 25,
        f'=IF(A{r}="","",'
        f'IF(OR(X{r}="Sanctioned",X{r}="Disbursed",X{r}="Approved"),"Approved",'
        f'IF(X{r}="Rejected","Rejected",'
        f'IF(OR(X{r}="Processing",X{r}="Documents Pending",X{r}="Vendor Assigned"),"In Progress",'
        f'IF(X{r}="Need Loan","Pending Start","—"))))')
    # Z DisbursementStatus
    hp.cell(r, 26,
        f'=IF(A{r}="","",IF(X{r}="Disbursed","Disbursed",'
        f'IF(X{r}="Sanctioned","Sanctioned - Awaiting Disbursement","Not Disbursed")))')
    # AA RiskCategory
    hp.cell(r, 27,
        f'=IF(A{r}="","",'
        f'IF(LOWER(TRIM({MD}!AI{r}&""))="yes","Need FLDG",'
        f'IF(LOWER(TRIM({MD}!AK{r}&""))="yes","Need Vidyalakshmi",'
        f'IF(LOWER(TRIM({MD}!AJ{r}&""))="yes","Need Review",'
        f'IF(LOWER(J{r})="critical","Critical",'
        f'IF(LOWER(U{r})="risk","Risk Case",'
        f'IF(X{r}="Refund","Refund","Normal")))))))')
    # AB CriticalFlag
    hp.cell(r, 28,
        f'=IF(A{r}="","",IF(OR(LOWER(J{r})="critical",LOWER(TRIM({MD}!AI{r}&""))="yes",'
        f'LOWER(TRIM({MD}!AK{r}&""))="yes",LOWER(U{r})="risk"),"Yes","No"))')
    # AC NeedFLDG
    hp.cell(r, 29, f'=IF(A{r}="","",IF(LOWER(TRIM({MD}!AI{r}&""))="yes","Yes","No"))')
    # AD NeedVidyalakshmi
    hp.cell(r, 30, f'=IF(A{r}="","",IF(LOWER(TRIM({MD}!AK{r}&""))="yes","Yes","No"))')
    # AE NeedVishwa
    hp.cell(r, 31, f'=IF(A{r}="","",IF(LOWER(TRIM({MD}!AJ{r}&""))="yes","Yes","No"))')
    # AF ReasonRaw
    hp.cell(r, 32, f'=IF(A{r}="","",TRIM({MD}!AL{r}&""))')
    # AG ReasonBucket — simplified keyword cascade
    hp.cell(r, 33,
        f'=IF(A{r}="","",IF(AF{r}="",—,IF(ISNUMBER(SEARCH("DNP",AF{r})),"Unreachable / DNP",'
        f'IF(ISNUMBER(SEARCH("interest",LOWER(AF{r}))),"High Interest / Terms",'
        f'IF(OR(ISNUMBER(SEARCH("eligib",LOWER(AF{r}))),ISNUMBER(SEARCH("income",LOWER(AF{r}))),ISNUMBER(SEARCH("Low",AF{r}))),"Eligibility / Income",'
        f'IF(OR(ISNUMBER(SEARCH("start",LOWER(AF{r}))),ISNUMBER(SEARCH("Newer",AF{r})),ISNUMBER(SEARCH("Fresh",AF{r})),ISNUMBER(SEARCH("Will start",AF{r}))),"Process Yet to Start",'
        f'IF(ISNUMBER(SEARCH("sure",LOWER(AF{r}))),"Undecided",'
        f'IF(OR(ISNUMBER(SEARCH("reject",LOWER(AF{r}))),ISNUMBER(SEARCH("SBI",AF{r}))),"Rejected Elsewhere",'
        f'IF(ISNUMBER(SEARCH("call",LOWER(AF{r}))),"Bank Communication",'
        f'IF(OR(ISNUMBER(SEARCH("father",LOWER(AF{r}))),ISNUMBER(SEARCH("sister",LOWER(AF{r}))),ISNUMBER(SEARCH("waiting",LOWER(AF{r})))),"Family / Dependency",'
        f'"Other"))))))))))')
    # Fix em-dash typo in formula above - use "-"
    
print('  (fixing ReasonBucket formula typo...)')

# Re-write ReasonBucket with proper dash
for r in range(2, N + 2):
    hp.cell(r, 33,
        f'=IF(A{r}="","",IF(AF{r}="","—",'
        f'IF(ISNUMBER(SEARCH("DNP",AF{r})),"Unreachable / DNP",'
        f'IF(ISNUMBER(SEARCH("interest",LOWER(AF{r}))),"High Interest / Terms",'
        f'IF(OR(ISNUMBER(SEARCH("eligib",LOWER(AF{r}))),ISNUMBER(SEARCH("income",LOWER(AF{r})))),"Eligibility / Income",'
        f'IF(OR(ISNUMBER(SEARCH("start",LOWER(AF{r}))),ISNUMBER(SEARCH("Newer",AF{r})),ISNUMBER(SEARCH("Will start",AF{r}))),"Process Yet to Start",'
        f'IF(ISNUMBER(SEARCH("sure",LOWER(AF{r}))),"Undecided",'
        f'IF(OR(ISNUMBER(SEARCH("reject",LOWER(AF{r}))),ISNUMBER(SEARCH("SBI",AF{r}))),"Rejected Elsewhere",'
        f'IF(ISNUMBER(SEARCH("call",LOWER(AF{r}))),"Bank Communication",'
        f'IF(OR(ISNUMBER(SEARCH("father",LOWER(AF{r}))),ISNUMBER(SEARCH("sister",LOWER(AF{r})))),"Family / Dependency",'
        f'"Other"))))))))))')
    # AH TentativeDate
    hp.cell(r, 34, f'=IF(A{r}="","",{MD}!AC{r})')
    # AI PendingDays
    hp.cell(r, 35,
        f'=IF(A{r}="","",IF(OR(AH{r}="",NOT(ISNUMBER(AH{r}))),"",'
        f'MAX(0,\'02_Configuration\'!$B$82-AH{r})))')
    # AJ AgeingBucket
    hp.cell(r, 36,
        f'=IF(A{r}="","",IF(AI{r}="","Unknown",'
        f'IF(AI{r}<=3,"0–3 Days",IF(AI{r}<=7,"4–7 Days",IF(AI{r}<=15,"8–15 Days",'
        f'IF(AI{r}<=30,"16–30 Days","30+ Days"))))))')
    # AK IsLoanPipeline
    hp.cell(r, 37, f'=IF(A{r}="","",IF(LOWER(H{r})="yes","Yes","No"))')
    # AL Pranjal
    hp.cell(r, 38, f'=IF(A{r}="","",{MD}!T{r})')
    # AM SST Remarks
    hp.cell(r, 39, f'=IF(A{r}="","",{MD}!S{r})')
    # AN LoanAmt
    hp.cell(r, 40, f'=IF(A{r}="","",{MD}!I{r})')
    # AO Income
    hp.cell(r, 41, f'=IF(A{r}="","",{MD}!M{r})')
    # AP-AT Processing flags from Loan Status
    hp.cell(r, 42, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("Icici",W{r}&"")),1,0))')
    hp.cell(r, 43, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("Propelld",W{r}&"")),1,0))')
    hp.cell(r, 44, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("GyanDhan",W{r}&"")),1,0))')
    hp.cell(r, 45, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("vidyalakshmi",LOWER(W{r}&""))),1,0))')
    hp.cell(r, 46, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("other bank",LOWER(W{r}&""))),1,0))')
    # AU IsSanctioned
    hp.cell(r, 47, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("sanctioned",LOWER(W{r}&""))),1,0))')
    # AV IsDocsPending
    hp.cell(r, 48, f'=IF(A{r}="","",IF(ISNUMBER(SEARCH("Docs Pending",W{r}&"")),1,0))')
    # AW IsRejectedStatus
    hp.cell(r, 49,
        f'=IF(A{r}="","",IF(OR(X{r}="Rejected",ISNUMBER(SEARCH("Not eligbile",W{r}&""))),1,0))')

hp.freeze_panes = 'C2'
hp.auto_filter.ref = f'A1:AW{N+1}'
for col in range(1, 50):
    hp.column_dimensions[get_column_letter(col)].width = 14
hp.column_dimensions['C'].width = 22
hp.column_dimensions['S'].width = 28
hp.column_dimensions['X'].width = 18
hp.column_dimensions['AL'].width = 24

print(f'  Helper formulas written for {N} rows')

# Save intermediate progress checkpoint by continuing in same script
print('Helper done. Continuing to analytics sheets...')

wb.save(OUT)
print(f'Checkpoint saved: {OUT}')
print('PASS1_DONE')
