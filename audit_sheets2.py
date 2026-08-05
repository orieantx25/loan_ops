# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'Master sheet - Loans .xlsx', data_only=False, read_only=False)

out = open('audit_sheets_out.txt', 'w', encoding='utf-8')

def p(s=''):
    out.write(s + '\n')

sheets = [
    'Data ', 'Data Summary', 'Dropdowns', 'Pivot Table 2',
    'Need FLDG attention', 'Student deferring loan requirem',
    'Need Vidyalakshmi Attention', 'Specific need students',
    'Students who want loan from nex', 'SummaryView', 'Sheet13', 'Phase 2',
    'ADYPU ', 'SSHAE', 'Refund Mapping', 'Merit data', 'Sheet18'
]

for sheet_name in sheets:
    if sheet_name not in wb.sheetnames:
        p(f'\nSKIP missing: {sheet_name}')
        continue
    ws = wb[sheet_name]
    p(f'\n========== [{sheet_name}] max_row={ws.max_row} max_col={ws.max_column} ==========')
    max_r = min(60, ws.max_row or 60)
    max_c = min(20, ws.max_column or 20)
    for row_idx in range(1, max_r + 1):
        cells = []
        for col_idx in range(1, max_c + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                v = str(cell.value).replace('\n', ' | ')[:100]
                cells.append(f'{get_column_letter(col_idx)}={v}')
        if cells:
            p(f'R{row_idx}: {" || ".join(cells)}')

p('\n========== DEFINED TABLES ==========')
for name in wb.sheetnames:
    ws = wb[name]
    if hasattr(ws, 'tables') and ws.tables:
        for tname, table in ws.tables.items():
            try:
                ref = getattr(table, 'ref', table)
                p(f'  Sheet[{name}] Table[{tname}] ref={ref}')
            except Exception as e:
                p(f'  Sheet[{name}] Table[{tname}] ERR {e}')

p('\n========== DEFINED NAMES ==========')
for dn in wb.defined_names.values():
    try:
        p(f'  {dn.name}: {dn.attr_text[:200]}')
    except Exception as e:
        p(f'  {dn.name}: ERR {e}')

# Count formula vs value cells per sheet
p('\n========== FORMULA DENSITY ==========')
for name in wb.sheetnames:
    ws = wb[name]
    formulas = 0
    values = 0
    blanks = 0
    # sample first 200 rows / 50 cols for speed
    for row in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row or 1), max_col=min(50, ws.max_column or 1)):
        for cell in row:
            if cell.value is None:
                blanks += 1
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                formulas += 1
            else:
                values += 1
    p(f'  [{name}] formulas~{formulas} values~{values} (sampled)')

wb.close()
out.close()
print('Wrote audit_sheets_out.txt')
