# -*- coding: utf-8 -*-
"""Pass 2: Student 360, Vendor/Stage/Risk analytics, Dashboard, Ops, Docs."""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

OUT = r'uGSOT_Loan_Operations_System.xlsx'
N = 492  # master rows
HLAST = N + 1

RED = 'E31C24'
BLACK = '111111'
DARK = '222222'
BG = 'F6F6F6'
WHITE = 'FFFFFF'
BORDER_C = 'E5E5E5'
MUTED = '6B6B6B'
LIGHT_RED = 'FDECEC'
GREEN = '1B7A4E'
AMBER = 'B45309'
GREY_HEADER = '2B2B2B'

thin = Border(
    left=Side(style='thin', color=BORDER_C),
    right=Side(style='thin', color=BORDER_C),
    top=Side(style='thin', color=BORDER_C),
    bottom=Side(style='thin', color=BORDER_C),
)
fill_red = PatternFill('solid', fgColor=RED)
fill_header = PatternFill('solid', fgColor=GREY_HEADER)
fill_bg = PatternFill('solid', fgColor=BG)
fill_white = PatternFill('solid', fgColor=WHITE)
fill_light_red = PatternFill('solid', fgColor=LIGHT_RED)
fill_black = PatternFill('solid', fgColor=BLACK)
fill_card = PatternFill('solid', fgColor=WHITE)

font_brand = Font(name='Calibri', size=22, bold=True, color=RED)
font_title = Font(name='Calibri', size=14, bold=True, color=BLACK)
font_h = Font(name='Calibri', size=10, bold=True, color=WHITE)
font_label = Font(name='Calibri', size=9, color=MUTED)
font_kpi = Font(name='Calibri', size=18, bold=True, color=BLACK)
font_kpi_red = Font(name='Calibri', size=18, bold=True, color=RED)
font_body = Font(name='Calibri', size=10, color=BLACK)
font_section = Font(name='Calibri', size=11, bold=True, color=BLACK)
font_small = Font(name='Calibri', size=8, color=MUTED)
font_filter = Font(name='Calibri', size=10, bold=True, color=BLACK)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

print('Opening workbook...')
wb = load_workbook(OUT)

# Remove sheets if re-run
for name in list(wb.sheetnames):
    if name not in ('01_Master_Data', '02_Configuration', '03_Helper'):
        del wb[name]

HP = "'03_Helper'"
CFG = "'02_Configuration'"
DT = "'08_Dashboard_Tables'"

def header_row(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row, start + i, h)
        cell.fill = fill_header
        cell.font = font_h
        cell.alignment = center
        cell.border = thin

def kpi_card(ws, r, c, label, formula, accent=False):
    """2-row KPI card: label + value."""
    label_cell = ws.cell(r, c, label)
    label_cell.font = font_label
    label_cell.fill = fill_white
    label_cell.alignment = center
    label_cell.border = thin
    val_cell = ws.cell(r + 1, c, formula)
    val_cell.font = font_kpi_red if accent else font_kpi
    val_cell.fill = fill_white
    val_cell.alignment = center
    val_cell.border = thin
    return val_cell

# ══════════════════════════════════════════════════════════════
# 04_Student_360
# ══════════════════════════════════════════════════════════════
print('Building 04_Student_360...')
s360 = wb.create_sheet('04_Student_360')
s360.sheet_properties.tabColor = '7C3AED'

# Because UNIQUE+XLOOKUP spill arrays work best in Google Sheets, we build
# Student 360 as a formula-driven table that mirrors Helper 1:1 for this dataset
# (Master is already ~1 student/row) PLUS explicit Unique metrics columns.
# A separate Unique Keys table powers vendor uniqueness.

s360_cols = [
    'StudentKey', 'StudentName', 'Campus', 'ProvisionalID', 'Mobile',
    'LoanRequired', 'CanonicalStage', 'ApprovalStatus', 'DisbursementStatus',
    'VendorCount', 'DuplicateVendorFlag', 'VendorsApplied', 'PrimaryVendor',
    'RiskCategory', 'CriticalFlag', 'NeedFLDG', 'NeedVidyalakshmi',
    'PendingDays', 'AgeingBucket', 'CaseStatus', 'LoanStatusRaw',
    'ReasonBucket', 'IsLoanPipeline', 'Scholarship', 'LoanAmt'
]
header_row(s360, 1, s360_cols)

# Map helper columns: B=Key C=Name G=Campus E=PID D=Mobile H=LoanReq X=Stage Y=Appr Z=Disb
# Q=VCount R=Dup S=Vendors T=Primary AA=Risk AB=Crit AC=FLDG AD=Vidya AI=Days AJ=Age
# U=Case W=Status AG=Reason AK=Pipeline K=Schol AN=Amt
helper_map = [
    'B', 'C', 'G', 'E', 'D', 'H', 'X', 'Y', 'Z',
    'Q', 'R', 'S', 'T', 'AA', 'AB', 'AC', 'AD',
    'AI', 'AJ', 'U', 'W', 'AG', 'AK', 'K', 'AN'
]

for r in range(2, HLAST + 1):
    for c, col in enumerate(helper_map, 1):
        s360.cell(r, c, f'=IF({HP}!A{r}="","",{HP}!{col}{r})')

s360.freeze_panes = 'C2'
s360.auto_filter.ref = f'A1:Y{HLAST}'
for col in range(1, 26):
    s360.column_dimensions[get_column_letter(col)].width = 14
s360.column_dimensions['B'].width = 22
s360.column_dimensions['L'].width = 26

# Unique key registry (for uniqueness proofs) — column AA onward
s360['AA1'] = 'UNIQUE_KEY_REGISTRY'
s360['AA1'].font = font_section
s360['AA2'] = 'Note: Master is already 1-row≈1-student. StudentKey enforces identity. Duplicate mobiles flagged below.'
s360['AA2'].font = font_small
s360['AA4'] = 'Duplicate Mobile Check'
s360['AA4'].font = font_section
s360['AA5'] = 'Mobile'
s360['AB5'] = 'Occurrences'
s360['AA5'].fill = fill_header
s360['AB5'].fill = fill_header
s360['AA5'].font = font_h
s360['AB5'].font = font_h
# Manual note — COUNTIF examples for the 2 known dups will surface via Ops
s360['AA6'] = '(Use Ops_Views → Duplicate Keys for live list)'
s360['AA6'].font = font_label

# ══════════════════════════════════════════════════════════════
# 05_Vendor_Analytics
# ══════════════════════════════════════════════════════════════
print('Building 05_Vendor_Analytics...')
va = wb.create_sheet('05_Vendor_Analytics')
va.sheet_properties.tabColor = '0F766E'

va['A1'] = 'VENDOR INTELLIGENCE'
va['A1'].font = font_brand
va.merge_cells('A1:L1')
va['A2'] = 'Unique Students = distinct students shared/processing with vendor. Applications = Shared-to Yes (or status token). Never sum Applications as student headcount.'
va['A2'].font = font_label
va.merge_cells('A2:L2')

vend_headers = [
    'Vendor', 'Active', 'Unique Students (Shared)', 'Applications (Shared Yes)',
    'In Loan Status / Processing', 'Sanctioned (status)', 'Docs Pending',
    'Pipeline Students (Latest=Yes ∩ vendor)', 'Critical Cases',
    'Approval %', 'Disbursement % (of shared)', 'Vendor Share % of Apps'
]
header_row(va, 4, vend_headers)

# Shared-flag vendors (rows 5-9) map to Helper L-P
shared_vendors = [
    ('ICICI', 'L', 'AP'),
    ('Propelld', 'M', 'AQ'),
    ('Study4Buddy', 'N', ''),
    ('Poonawala Fincorp', 'O', ''),
    ('GyanDhan', 'P', 'AR'),
]
# Status-only vendors
status_vendors = [
    ('PM Vidyalakshmi', '', 'AS'),
    ('Other Bank', '', 'AT'),
]

row = 5
for name, share_col, status_col in shared_vendors:
    va.cell(row, 1, name).border = thin
    va.cell(row, 2, 'Yes').border = thin
    # Unique Students Shared = COUNTIFS where share=1 (1 row/student already)
    va.cell(row, 3, f'=COUNTIF({HP}!{share_col}:{share_col},1)').border = thin
    va.cell(row, 4, f'=C{row}').border = thin  # apps = same when 1 row/student
    if status_col:
        va.cell(row, 5, f'=COUNTIF({HP}!{status_col}:{status_col},1)').border = thin
    else:
        va.cell(row, 5, 0).border = thin
    va.cell(row, 6,
        f'=COUNTIFS({HP}!{share_col}:{share_col},1,{HP}!AU:AU,1)').border = thin
    va.cell(row, 7,
        f'=COUNTIFS({HP}!{share_col}:{share_col},1,{HP}!AV:AV,1)').border = thin
    va.cell(row, 8,
        f'=COUNTIFS({HP}!{share_col}:{share_col},1,{HP}!AK:AK,"Yes")').border = thin
    va.cell(row, 9,
        f'=COUNTIFS({HP}!{share_col}:{share_col},1,{HP}!AB:AB,"Yes")').border = thin
    # Approval % = sanctioned / shared
    va.cell(row, 10, f'=IF(C{row}=0,0,F{row}/C{row})').border = thin
    va.cell(row, 10).number_format = '0.0%'
    # Disbursement % among shared
    va.cell(row, 11,
        f'=IF(C{row}=0,0,COUNTIFS({HP}!{share_col}:{share_col},1,{HP}!X:X,"Disbursed")/C{row})').border = thin
    va.cell(row, 11).number_format = '0.0%'
    va.cell(row, 12, f'=IF(SUM($D$5:$D$9)=0,0,D{row}/SUM($D$5:$D$9))').border = thin
    va.cell(row, 12).number_format = '0.0%'
    row += 1

for name, share_col, status_col in status_vendors:
    va.cell(row, 1, name).border = thin
    va.cell(row, 2, 'Yes').border = thin
    va.cell(row, 3, f'=COUNTIF({HP}!{status_col}:{status_col},1)').border = thin
    va.cell(row, 4, f'=C{row}').border = thin
    va.cell(row, 5, f'=C{row}').border = thin
    va.cell(row, 6, f'=COUNTIFS({HP}!{status_col}:{status_col},1,{HP}!AU:AU,1)').border = thin
    va.cell(row, 7, f'=COUNTIFS({HP}!{status_col}:{status_col},1,{HP}!AV:AV,1)').border = thin
    va.cell(row, 8, f'=COUNTIFS({HP}!{status_col}:{status_col},1,{HP}!AK:AK,"Yes")').border = thin
    va.cell(row, 9, f'=COUNTIFS({HP}!{status_col}:{status_col},1,{HP}!AB:AB,"Yes")').border = thin
    va.cell(row, 10, f'=IF(C{row}=0,0,F{row}/C{row})').border = thin
    va.cell(row, 10).number_format = '0.0%'
    va.cell(row, 11, f'=IF(C{row}=0,0,COUNTIFS({HP}!{status_col}:{status_col},1,{HP}!X:X,"Disbursed")/C{row})').border = thin
    va.cell(row, 11).number_format = '0.0%'
    va.cell(row, 12, '—').border = thin
    row += 1

# Multi-vendor distribution
va['A14'] = 'MULTI-VENDOR DISTRIBUTION (Unique Students)'
va['A14'].font = font_section
va['A14'].fill = fill_light_red
header_row(va, 15, ['Vendor Count', 'Students', '% of Shared Students'])
for i, label in enumerate(['0 Vendors', '1 Vendor', '2 Vendors', '3 Vendors', '4 Vendors', '5+ Vendors'], 16):
    va.cell(i, 1, label).border = thin
    if i < 21:
        cnt = i - 16
        va.cell(i, 2, f'=COUNTIF({HP}!Q:Q,{cnt})').border = thin
    else:
        va.cell(i, 2, f'=COUNTIFS({HP}!Q:Q,">=5")').border = thin
    va.cell(i, 3, f'=IF(COUNTA({HP}!A:A)-1=0,0,B{i}/(COUNTA({HP}!A:A)-1))').border = thin
    va.cell(i, 3).number_format = '0.0%'

va['A24'] = 'SUMMARY'
va['A24'].font = font_section
va['A25'] = 'Students with ≥1 vendor'
va['B25'] = f'=COUNTIFS({HP}!Q:Q,">=1")'
va['A26'] = 'Students with duplicate vendors (≥2)'
va['B26'] = f'=COUNTIFS({HP}!Q:Q,">=2")'
va['A27'] = 'Total vendor applications (Shared Yes sum)'
va['B27'] = '=SUM(D5:D9)'
va['A28'] = 'Avg vendors per shared student'
va['B28'] = '=IF(B25=0,0,B27/B25)'
va['B28'].number_format = '0.00'

# Vendor Overlap Matrix
va['A31'] = 'VENDOR OVERLAP MATRIX (students shared to both)'
va['A31'].font = font_section
va['A31'].fill = fill_light_red
overlap_names = ['ICICI', 'Propelld', 'Study4Buddy', 'Poonawala', 'GyanDhan']
overlap_cols = ['L', 'M', 'N', 'O', 'P']
va.cell(32, 1, 'Vendor').fill = fill_header
va.cell(32, 1).font = font_h
for i, name in enumerate(overlap_names):
    cell = va.cell(32, i + 2, name)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin
    cell.alignment = center
    cell = va.cell(33 + i, 1, name)
    cell.fill = fill_header
    cell.font = font_h
    cell.border = thin

for i, c1 in enumerate(overlap_cols):
    for j, c2 in enumerate(overlap_cols):
        cell = va.cell(33 + i, 2 + j)
        cell.border = thin
        cell.alignment = center
        if i == j:
            cell.value = '—'
            cell.fill = PatternFill('solid', fgColor='EEEEEE')
        else:
            cell.value = f'=COUNTIFS({HP}!{c1}:{c1},1,{HP}!{c2}:{c2},1)'

# Color scale on overlap
va.conditional_formatting.add('B33:F37', ColorScaleRule(
    start_type='num', start_value=0, start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color='FECACA',
    end_type='max', end_color=RED
))

for col in range(1, 13):
    va.column_dimensions[get_column_letter(col)].width = 14
va.column_dimensions['A'].width = 28

# ══════════════════════════════════════════════════════════════
# 06_Stage_Analytics
# ══════════════════════════════════════════════════════════════
print('Building 06_Stage_Analytics...')
sa = wb.create_sheet('06_Stage_Analytics')
sa.sheet_properties.tabColor = 'B45309'

sa['A1'] = 'STAGE ANALYTICS & LOAN FUNNEL'
sa['A1'].font = font_brand
sa['A2'] = 'Every student has exactly one CanonicalStage (see 03_Helper). Funnel Core stages drive executive conversion.'
sa['A2'].font = font_label

header_row(sa, 4, ['Stage', 'FunnelOrder', 'Students', '% of Total', '% of Pipeline (Latest=Yes)', 'Conversion from Prior Funnel'])

funnel_stages = [
    ('Need Loan', 1),
    ('Interested', 2),
    ('Loan Started', 3),
    ('Documents Pending', 4),
    ('Vendor Assigned', 5),
    ('Processing', 6),
    ('Approved', 7),
    ('Sanctioned', 8),
    ('Disbursed', 9),
    ('Completed', 10),
    ('Rejected', 90),
    ('Refund', 91),
    ('DNP', 92),
    ('Not Required', 93),
    ('Unclassified', 99),
]

for i, (stage, order) in enumerate(funnel_stages):
    r = 5 + i
    sa.cell(r, 1, stage).border = thin
    sa.cell(r, 2, order).border = thin
    sa.cell(r, 3, f'=COUNTIF({HP}!X:X,A{r})').border = thin
    sa.cell(r, 4, f'=IF(COUNTA({HP}!A:A)<=1,0,C{r}/(COUNTA({HP}!A:A)-1))').border = thin
    sa.cell(r, 4).number_format = '0.0%'
    sa.cell(r, 5, f'=IF(COUNTIF({HP}!AK:AK,"Yes")=0,0,COUNTIFS({HP}!X:X,A{r},{HP}!AK:AK,"Yes")/COUNTIF({HP}!AK:AK,"Yes"))').border = thin
    sa.cell(r, 5).number_format = '0.0%'
    if i == 0:
        sa.cell(r, 6, '—').border = thin
    else:
        sa.cell(r, 6, f'=IF(C{r-1}=0,"—",C{r}/C{r-1})').border = thin
        sa.cell(r, 6).number_format = '0.0%'

# Executive funnel (core only)
sa['A22'] = 'EXECUTIVE FUNNEL (Core)'
sa['A22'].font = font_section
sa['A22'].fill = fill_light_red
header_row(sa, 23, ['Funnel Stage', 'Count', 'Conversion vs Need Loan', 'Step Conversion'])

core = ['Need Loan', 'Loan Started', 'Documents Pending', 'Processing', 'Sanctioned', 'Disbursed']
# Note: Loan Started in our mapping is fallback; Processing is main body.
# For exec funnel use operational definitions:
# Need Loan = Latest Yes & stage Need Loan
# Started = Latest Yes (all pipeline) OR stage in started+
# We'll define explicit funnel metrics:

sa['A24'] = 'Need Loan (Latest=Yes)'
sa['B24'] = f'=COUNTIF({HP}!AK:AK,"Yes")'
sa['C24'] = '100%'
sa['D24'] = '—'

sa['A25'] = 'Not Started'
sa['B25'] = f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Need Loan")+COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Vendor Assigned")'
sa['C25'] = '=IF(B24=0,0,B25/B24)'
sa['C25'].number_format = '0.0%'
sa['D25'] = '—'

sa['A26'] = 'Processing'
sa['B26'] = f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Processing")+COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Documents Pending")'
sa['C26'] = '=IF(B24=0,0,B26/B24)'
sa['C26'].number_format = '0.0%'
sa['D26'] = '=IF(B25=0,0,B26/B25)'
sa['D26'].number_format = '0.0%'

sa['A27'] = 'Sanctioned / Approved'
sa['B27'] = f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Sanctioned")+COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Approved")'
sa['C27'] = '=IF(B24=0,0,B27/B24)'
sa['C27'].number_format = '0.0%'
sa['D27'] = '=IF(B26=0,0,B27/B26)'
sa['D27'].number_format = '0.0%'

sa['A28'] = 'Disbursed'
sa['B28'] = f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Disbursed")'
sa['C28'] = '=IF(B24=0,0,B28/B24)'
sa['C28'].number_format = '0.0%'
sa['D28'] = '=IF(B27=0,0,B28/B27)'
sa['D28'].number_format = '0.0%'

sa['A29'] = 'Rejected'
sa['B29'] = f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Rejected")'
sa['C29'] = '=IF(B24=0,0,B29/B24)'
sa['C29'].number_format = '0.0%'

for r in range(24, 30):
    for c in range(1, 5):
        sa.cell(r, c).border = thin
        sa.cell(r, c).font = font_body

sa.column_dimensions['A'].width = 36
sa.column_dimensions['B'].width = 12
sa.column_dimensions['C'].width = 22
sa.column_dimensions['D'].width = 18
sa.column_dimensions['E'].width = 24
sa.column_dimensions['F'].width = 22

# ══════════════════════════════════════════════════════════════
# 07_Risk_Analytics
# ══════════════════════════════════════════════════════════════
print('Building 07_Risk_Analytics...')
ra = wb.create_sheet('07_Risk_Analytics')
ra.sheet_properties.tabColor = RED

ra['A1'] = 'RISK & ATTENTION ANALYTICS'
ra['A1'].font = font_brand

header_row(ra, 3, ['Risk Flag', 'Count', '% of Pipeline', '% of All Students'])

risk_rows = [
    ('Need FLDG', f'=COUNTIF({HP}!AC:AC,"Yes")'),
    ('Need Vidyalakshmi', f'=COUNTIF({HP}!AD:AD,"Yes")'),
    ('Need Vishwa Review', f'=COUNTIF({HP}!AE:AE,"Yes")'),
    ('Critical Flag', f'=COUNTIF({HP}!AB:AB,"Yes")'),
    ('Risk Case Status', f'=COUNTIF({HP}!U:U,"Risk")'),
    ('Refund', f'=COUNTIF({HP}!X:X,"Refund")'),
    ('DNP / Unreachable', f'=COUNTIF({HP}!X:X,"DNP")'),
    ('Rejected', f'=COUNTIF({HP}!X:X,"Rejected")'),
    ('Documents Pending', f'=COUNTIF({HP}!X:X,"Documents Pending")'),
    ('Duplicate Vendor Cases', f'=COUNTIF({HP}!R:R,"Yes")'),
    ('Not Started (Pipeline)', f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Need Loan")'),
]

for i, (label, formula) in enumerate(risk_rows):
    r = 4 + i
    ra.cell(r, 1, label).border = thin
    ra.cell(r, 2, formula).border = thin
    ra.cell(r, 3, f'=IF(COUNTIF({HP}!AK:AK,"Yes")=0,0,B{r}/COUNTIF({HP}!AK:AK,"Yes"))').border = thin
    ra.cell(r, 3).number_format = '0.0%'
    ra.cell(r, 4, f'=IF(COUNTA({HP}!A:A)<=1,0,B{r}/(COUNTA({HP}!A:A)-1))').border = thin
    ra.cell(r, 4).number_format = '0.0%'

ra['A17'] = 'REASON ANALYSIS (Not Started)'
ra['A17'].font = font_section
ra['A17'].fill = fill_light_red
header_row(ra, 18, ['Reason Bucket', 'Count'])
reason_buckets = [
    'Unreachable / DNP', 'High Interest / Terms', 'Eligibility / Income',
    'Process Yet to Start', 'Undecided', 'Rejected Elsewhere',
    'Bank Communication', 'Family / Dependency', 'Other', '—'
]
for i, b in enumerate(reason_buckets):
    r = 19 + i
    ra.cell(r, 1, b).border = thin
    ra.cell(r, 2, f'=COUNTIF({HP}!AG:AG,A{r})').border = thin

ra['A32'] = 'AGEING DISTRIBUTION (students with Tentative Date)'
ra['A32'].font = font_section
header_row(ra, 33, ['Ageing Bucket', 'Count'])
for i, b in enumerate(['0–3 Days', '4–7 Days', '8–15 Days', '16–30 Days', '30+ Days', 'Unknown']):
    r = 34 + i
    ra.cell(r, 1, b).border = thin
    ra.cell(r, 2, f'=COUNTIF({HP}!AJ:AJ,A{r})').border = thin

ra.column_dimensions['A'].width = 32
ra.column_dimensions['B'].width = 12
ra.column_dimensions['C'].width = 14
ra.column_dimensions['D'].width = 16

# ══════════════════════════════════════════════════════════════
# 08_Dashboard_Tables  (all KPI sources — Dashboard only references these)
# ══════════════════════════════════════════════════════════════
print('Building 08_Dashboard_Tables...')
dt = wb.create_sheet('08_Dashboard_Tables')
dt.sheet_properties.tabColor = '333333'

dt['A1'] = 'DASHBOARD DATA TABLES — Do not format for presentation. Dashboard reads from here.'
dt['A1'].font = font_label

# Filters (mirrored from Dashboard — Dashboard writes, Tables reads)
dt['A3'] = 'ACTIVE FILTERS'
dt['A3'].font = font_section
dt['A4'] = 'Campus'
dt['B4'] = '(All)'  # Dashboard will link here; default All
dt['A5'] = 'Vendor'
dt['B5'] = '(All)'
dt['A6'] = 'Stage'
dt['B6'] = '(All)'
dt['A7'] = 'Critical'
dt['B7'] = '(All)'
dt['A8'] = 'LoanRequired'
dt['B8'] = '(All)'
dt['A9'] = 'DuplicateVendor'
dt['B9'] = '(All)'

# Note: Full dynamic FILTER across all KPIs in pure Sheets needs FILTER wrapper.
# For v1 we compute unfiltered global KPIs (instant leadership view) and
# provide filtered Student list on Ops_Views. Filter cells reserved for Sheets enhancement.

dt['A11'] = 'EXECUTIVE KPIs'
dt['A11'].font = font_section
dt['A11'].fill = fill_light_red

kpis = [
    ('Total Students', f'=COUNTA({HP}!A:A)-1'),
    ('Need Loan (Latest=Yes)', f'=COUNTIF({HP}!AK:AK,"Yes")'),
    ('Loan Started / In Motion', f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"<>Need Loan")-COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Not Required")'),
    ('Processing', f'=COUNTIF({HP}!X:X,"Processing")'),
    ('Documents Pending', f'=COUNTIF({HP}!X:X,"Documents Pending")+COUNTIF({HP}!AV:AV,1)'),
    ('Sanctioned', f'=COUNTIF({HP}!X:X,"Sanctioned")'),
    ('Disbursed', f'=COUNTIF({HP}!X:X,"Disbursed")'),
    ('Rejected', f'=COUNTIF({HP}!X:X,"Rejected")'),
    ('Need FLDG', f'=COUNTIF({HP}!AC:AC,"Yes")'),
    ('Need Vidyalakshmi', f'=COUNTIF({HP}!AD:AD,"Yes")'),
    ('Critical Cases', f'=COUNTIF({HP}!AB:AB,"Yes")'),
    ('Duplicate Vendor Students', f'=COUNTIF({HP}!R:R,"Yes")'),
    ('Avg Vendors (shared students)', f"='05_Vendor_Analytics'!B28"),
    ('Students with ≥1 Vendor', f"='05_Vendor_Analytics'!B25"),
    ('Not Started (Pipeline)', f'=COUNTIFS({HP}!AK:AK,"Yes",{HP}!X:X,"Need Loan")'),
    ('Refund', f'=COUNTIF({HP}!X:X,"Refund")'),
    ('DNP', f'=COUNTIF({HP}!X:X,"DNP")'),
    ('Not Required', f'=COUNTIF({HP}!X:X,"Not Required")'),
    ('Control Cases', f'=COUNTIF({HP}!U:U,"Control")'),
    ('Risk Cases', f'=COUNTIF({HP}!U:U,"Risk")'),
]

dt['A12'] = 'KPI'
dt['B12'] = 'Value'
dt['A12'].fill = fill_header
dt['B12'].fill = fill_header
dt['A12'].font = font_h
dt['B12'].font = font_h

for i, (label, formula) in enumerate(kpis):
    r = 13 + i
    dt.cell(r, 1, label).border = thin
    dt.cell(r, 2, formula).border = thin
    dt.cell(r, 2).font = Font(name='Calibri', size=12, bold=True, color=BLACK)

# Funnel table for dashboard
dt['D11'] = 'FUNNEL'
dt['D11'].font = font_section
header_row(dt, 12, ['Stage', 'Count', 'Conv%'], start=4)
for i, (label, src) in enumerate([
    ('Need Loan', "='06_Stage_Analytics'!B24"),
    ('Not Started', "='06_Stage_Analytics'!B25"),
    ('Processing', "='06_Stage_Analytics'!B26"),
    ('Sanctioned', "='06_Stage_Analytics'!B27"),
    ('Disbursed', "='06_Stage_Analytics'!B28"),
    ('Rejected', "='06_Stage_Analytics'!B29"),
]):
    r = 13 + i
    dt.cell(r, 4, label).border = thin
    dt.cell(r, 5, src).border = thin
    if i == 0:
        dt.cell(r, 6, 1).border = thin
    else:
        dt.cell(r, 6, f'=IF($E$13=0,0,E{r}/$E$13)').border = thin
    dt.cell(r, 6).number_format = '0.0%'

# Campus table
dt['H11'] = 'CAMPUS DASHBOARD'
dt['H11'].font = font_section
header_row(dt, 12, ['Campus', 'Total', 'Need Loan', 'Processing', 'Sanctioned', 'Disbursed', 'Rejected', 'Critical'], start=8)
for i, campus in enumerate(['SSAHE', 'ADYPU']):
    r = 13 + i
    dt.cell(r, 8, campus).border = thin
    dt.cell(r, 9, f'=COUNTIF({HP}!G:G,H{r})').border = thin
    dt.cell(r, 10, f'=COUNTIFS({HP}!G:G,H{r},{HP}!AK:AK,"Yes")').border = thin
    dt.cell(r, 11, f'=COUNTIFS({HP}!G:G,H{r},{HP}!X:X,"Processing")').border = thin
    dt.cell(r, 12, f'=COUNTIFS({HP}!G:G,H{r},{HP}!X:X,"Sanctioned")').border = thin
    dt.cell(r, 13, f'=COUNTIFS({HP}!G:G,H{r},{HP}!X:X,"Disbursed")').border = thin
    dt.cell(r, 14, f'=COUNTIFS({HP}!G:G,H{r},{HP}!X:X,"Rejected")').border = thin
    dt.cell(r, 15, f'=COUNTIFS({HP}!G:G,H{r},{HP}!AB:AB,"Yes")').border = thin

# Vendor scorecard compact
dt['A36'] = 'VENDOR SCORECARD'
dt['A36'].font = font_section
header_row(dt, 37, ['Vendor', 'Unique', 'Apps', 'Processing', 'Sanctioned', 'Critical', 'Approval%'])
for i in range(5):
    r = 38 + i
    src = 5 + i
    dt.cell(r, 1, f"='05_Vendor_Analytics'!A{src}").border = thin
    dt.cell(r, 2, f"='05_Vendor_Analytics'!C{src}").border = thin
    dt.cell(r, 3, f"='05_Vendor_Analytics'!D{src}").border = thin
    dt.cell(r, 4, f"='05_Vendor_Analytics'!E{src}").border = thin
    dt.cell(r, 5, f"='05_Vendor_Analytics'!F{src}").border = thin
    dt.cell(r, 6, f"='05_Vendor_Analytics'!I{src}").border = thin
    dt.cell(r, 7, f"='05_Vendor_Analytics'!J{src}").border = thin
    dt.cell(r, 7).number_format = '0.0%'

# Risk compact
dt['A46'] = 'RISK STRIP'
header_row(dt, 47, ['Flag', 'Count'])
for i in range(10):
    r = 48 + i
    dt.cell(r, 1, f"='07_Risk_Analytics'!A{4+i}").border = thin
    dt.cell(r, 2, f"='07_Risk_Analytics'!B{4+i}").border = thin

# Ageing
dt['D46'] = 'AGEING'
header_row(dt, 47, ['Bucket', 'Count'], start=4)
for i in range(6):
    r = 48 + i
    dt.cell(r, 4, f"='07_Risk_Analytics'!A{34+i}").border = thin
    dt.cell(r, 5, f"='07_Risk_Analytics'!B{34+i}").border = thin

# Reason
dt['G46'] = 'REASONS'
header_row(dt, 47, ['Bucket', 'Count'], start=7)
for i in range(9):
    r = 48 + i
    dt.cell(r, 7, f"='07_Risk_Analytics'!A{19+i}").border = thin
    dt.cell(r, 8, f"='07_Risk_Analytics'!B{19+i}").border = thin

# Loan status distribution (canonical)
dt['A60'] = 'STAGE DISTRIBUTION'
header_row(dt, 61, ['Stage', 'Count'])
for i, stage in enumerate(['Need Loan', 'Vendor Assigned', 'Documents Pending', 'Processing',
                            'Sanctioned', 'Disbursed', 'Rejected', 'Refund', 'DNP', 'Not Required', 'Interested', 'Unclassified']):
    r = 62 + i
    dt.cell(r, 1, stage).border = thin
    dt.cell(r, 2, f'=COUNTIF({HP}!X:X,A{r})').border = thin

dt.column_dimensions['A'].width = 32
dt.column_dimensions['B'].width = 14
dt.column_dimensions['D'].width = 16
dt.column_dimensions['H'].width = 12

# ══════════════════════════════════════════════════════════════
# 09_Dashboard — Executive UI
# ══════════════════════════════════════════════════════════════
print('Building 09_Dashboard...')
db = wb.create_sheet('09_Dashboard', 0)  # first sheet
db.sheet_properties.tabColor = RED

# Background wash
for r in range(1, 70):
    for c in range(1, 16):
        db.cell(r, c).fill = fill_bg

# Header bar
for c in range(1, 15):
    db.cell(1, c).fill = fill_black
    db.cell(2, c).fill = fill_black
db['A1'] = 'upGrad School of Technology'
db['A1'].font = Font(name='Calibri', size=11, bold=True, color=RED)
db.merge_cells('A1:D1')
db['A2'] = 'LOAN OPERATIONS MANAGEMENT SYSTEM'
db['A2'].font = Font(name='Calibri', size=18, bold=True, color=WHITE)
db.merge_cells('A2:H2')
db['I1'] = 'Admission Cycle'
db['I1'].font = Font(name='Calibri', size=8, color='AAAAAA')
db['I2'] = f"={CFG}!B83"
db['I2'].font = Font(name='Calibri', size=11, bold=True, color=WHITE)
db['K1'] = 'As of'
db['K1'].font = Font(name='Calibri', size=8, color='AAAAAA')
db['K2'] = f"={CFG}!B82"
db['K2'].font = Font(name='Calibri', size=11, bold=True, color=WHITE)
db['K2'].number_format = 'DD-MMM-YYYY'
db['M1'] = 'Auto-updating'
db['M1'].font = Font(name='Calibri', size=8, color='AAAAAA')
db['M2'] = 'Formula-driven'
db['M2'].font = Font(name='Calibri', size=10, bold=True, color=RED)

db.row_dimensions[1].height = 18
db.row_dimensions[2].height = 28

# Filter strip
db['A4'] = 'FILTERS'
db['A4'].font = font_section
db['A4'].fill = fill_light_red
filters = [
    (5, 'Campus', 'B4', ['(All)', 'SSAHE', 'ADYPU']),
    (6, 'Critical', 'B7', ['(All)', 'Yes', 'No']),
    (7, 'Loan Required', 'B8', ['(All)', 'Yes', 'No', 'Refund', 'DNP']),
    (8, 'Duplicate Vendor', 'B9', ['(All)', 'Yes', 'No']),
]
db['B4'] = 'Campus'
db['C4'] = "='08_Dashboard_Tables'!B4"
db['D4'] = 'Critical'
db['E4'] = "='08_Dashboard_Tables'!B7"
db['F4'] = 'Loan Required'
db['G4'] = "='08_Dashboard_Tables'!B8"
db['H4'] = 'Dup Vendor'
db['I4'] = "='08_Dashboard_Tables'!B9"
for col in ['B', 'D', 'F', 'H']:
    db[f'{col}4'].font = font_filter
for col in ['C', 'E', 'G', 'I']:
    db[f'{col}4'].fill = fill_white
    db[f'{col}4'].border = thin
    db[f'{col}4'].font = font_body

db['K4'] = 'Edit filters on 08_Dashboard_Tables!B4:B9  →  Ops lists honor them in v2. KPI strip = global truth.'
db['K4'].font = font_small

# Section A — KPI cards
db['A6'] = 'EXECUTIVE SNAPSHOT'
db['A6'].font = font_title
db['A6'].fill = fill_bg

kpi_layout = [
    # row 7-8
    (7, 1, 'Total Students', "='08_Dashboard_Tables'!B13", False),
    (7, 2, 'Need Loan', "='08_Dashboard_Tables'!B14", True),
    (7, 3, 'Processing', "='08_Dashboard_Tables'!B16", False),
    (7, 4, 'Docs Pending', "='08_Dashboard_Tables'!B17", False),
    (7, 5, 'Sanctioned', "='08_Dashboard_Tables'!B18", False),
    (7, 6, 'Disbursed', "='08_Dashboard_Tables'!B19", False),
    (7, 7, 'Rejected', "='08_Dashboard_Tables'!B20", True),
    # row 10-11
    (10, 1, 'Need FLDG', "='08_Dashboard_Tables'!B21", True),
    (10, 2, 'Need Vidyalakshmi', "='08_Dashboard_Tables'!B22", True),
    (10, 3, 'Critical Cases', "='08_Dashboard_Tables'!B23", True),
    (10, 4, 'Dup Vendors', "='08_Dashboard_Tables'!B24", False),
    (10, 5, 'Avg Vendors', "='08_Dashboard_Tables'!B25", False),
    (10, 6, 'With Vendor', "='08_Dashboard_Tables'!B26", False),
    (10, 7, 'Not Started', "='08_Dashboard_Tables'!B27", False),
]

for r, c, label, formula, accent in kpi_layout:
    kpi_card(db, r, c, label, formula, accent)
    db.column_dimensions[get_column_letter(c)].width = 14

db.row_dimensions[8].height = 32
db.row_dimensions[11].height = 32

# Section B — Funnel
db['A13'] = 'LOAN FUNNEL'
db['A13'].font = font_title

header_row(db, 14, ['Stage', 'Students', 'vs Need Loan'])
for i in range(6):
    r = 15 + i
    db.cell(r, 1, f"='08_Dashboard_Tables'!D{13+i}").border = thin
    db.cell(r, 1).fill = fill_white
    db.cell(r, 2, f"='08_Dashboard_Tables'!E{13+i}").border = thin
    db.cell(r, 2).fill = fill_white
    db.cell(r, 2).font = font_kpi if i == 0 else font_body
    db.cell(r, 3, f"='08_Dashboard_Tables'!F{13+i}").border = thin
    db.cell(r, 3).fill = fill_white
    db.cell(r, 3).number_format = '0.0%'

# Visual funnel bars via conditional width text (simple bar column)
db['D14'] = 'Bar'
db['D14'].fill = fill_header
db['D14'].font = font_h
for i in range(6):
    r = 15 + i
    # REPT bar proportional
    db.cell(r, 4, f'=IF($B$15=0,"",REPT("█",ROUND(B{r}/$B$15*20,0)))').border = thin
    db.cell(r, 4).font = Font(name='Calibri', size=10, color=RED)
    db.cell(r, 4).fill = fill_white

# Section C — Vendor Performance
db['F13'] = 'VENDOR PERFORMANCE'
db['F13'].font = font_title
header_row(db, 14, ['Vendor', 'Unique', 'Apps', 'In Status', 'Sanctioned', 'Critical', 'Appr%'], start=6)
for i in range(5):
    r = 15 + i
    for c in range(7):
        db.cell(r, 6 + c, f"='08_Dashboard_Tables'!{get_column_letter(1+c)}{38+i}").border = thin
        db.cell(r, 6 + c).fill = fill_white
    db.cell(r, 12).number_format = '0.0%'

db['F21'] = 'Overlap matrix → see 05_Vendor_Analytics (heatmap)'
db['F21'].font = font_small

# Section D — Campus
db['A22'] = 'CAMPUS PERFORMANCE'
db['A22'].font = font_title
header_row(db, 23, ['Campus', 'Total', 'Need Loan', 'Processing', 'Sanctioned', 'Disbursed', 'Rejected', 'Critical'])
for i in range(2):
    r = 24 + i
    for c in range(8):
        db.cell(r, 1 + c, f"='08_Dashboard_Tables'!{get_column_letter(8+c)}{13+i}").border = thin
        db.cell(r, 1 + c).fill = fill_white

# Section E — Risk
db['A27'] = 'RISK DASHBOARD'
db['A27'].font = font_title
header_row(db, 28, ['Flag', 'Count'])
for i in range(10):
    r = 29 + i
    db.cell(r, 1, f"='08_Dashboard_Tables'!A{48+i}").border = thin
    db.cell(r, 1).fill = fill_white
    db.cell(r, 2, f"='08_Dashboard_Tables'!B{48+i}").border = thin
    db.cell(r, 2).fill = fill_white
    if i < 3:
        db.cell(r, 2).font = Font(name='Calibri', size=12, bold=True, color=RED)

# Ageing
db['D27'] = 'STUDENT AGEING'
db['D27'].font = font_title
header_row(db, 28, ['Bucket', 'Count'], start=4)
for i in range(6):
    r = 29 + i
    db.cell(r, 4, f"='08_Dashboard_Tables'!D{48+i}").border = thin
    db.cell(r, 4).fill = fill_white
    db.cell(r, 5, f"='08_Dashboard_Tables'!E{48+i}").border = thin
    db.cell(r, 5).fill = fill_white
    db.cell(r, 6, f'=IF(SUM($E$29:$E$34)=0,"",REPT("█",ROUND(E{r}/MAX($E$29:$E$34)*15,0)))').border = thin
    db.cell(r, 6).font = Font(name='Calibri', size=9, color=DARK)
    db.cell(r, 6).fill = fill_white

# Reasons
db['H27'] = 'DROP-OFF REASONS'
db['H27'].font = font_title
header_row(db, 28, ['Reason', 'Count'], start=8)
for i in range(9):
    r = 29 + i
    db.cell(r, 8, f"='08_Dashboard_Tables'!G{48+i}").border = thin
    db.cell(r, 8).fill = fill_white
    db.cell(r, 9, f"='08_Dashboard_Tables'!H{48+i}").border = thin
    db.cell(r, 9).fill = fill_white

# Stage distribution
db['A41'] = 'LOAN STATUS DISTRIBUTION'
db['A41'].font = font_title
header_row(db, 42, ['Stage', 'Count', 'Share'])
for i in range(12):
    r = 43 + i
    db.cell(r, 1, f"='08_Dashboard_Tables'!A{62+i}").border = thin
    db.cell(r, 1).fill = fill_white
    db.cell(r, 2, f"='08_Dashboard_Tables'!B{62+i}").border = thin
    db.cell(r, 2).fill = fill_white
    db.cell(r, 3, f'=IF(SUM($B$43:$B$54)=0,0,B{r}/SUM($B$43:$B$54))').border = thin
    db.cell(r, 3).number_format = '0.0%'
    db.cell(r, 3).fill = fill_white
    db.cell(r, 4, f'=IF(SUM($B$43:$B$54)=0,"",REPT("█",ROUND(B{r}/MAX($B$43:$B$54)*18,0)))').border = thin
    db.cell(r, 4).font = Font(name='Calibri', size=9, color=RED)
    db.cell(r, 4).fill = fill_white

# Top pending pointer
db['F41'] = 'TOP PENDING STUDENTS'
db['F41'].font = font_title
db['F42'] = 'See sheet 10_Ops_Views for live Top Pending, FLDG list, Vidyalakshmi list, and student search.'
db['F42'].font = font_label
db.merge_cells('F42:L42')

db['F44'] = 'QUICK NAVIGATION'
db['F44'].font = font_section
nav = [
    ('F45', '→ 04_Student_360 — One row per student'),
    ('F46', '→ 05_Vendor_Analytics — Unique vs Apps + Overlap heatmap'),
    ('F47', '→ 06_Stage_Analytics — Full funnel'),
    ('F48', '→ 07_Risk_Analytics — Risk & reasons'),
    ('F49', '→ 10_Ops_Views — Action lists'),
    ('F50', '→ 01_Master_Data — Edit data here ONLY'),
    ('F51', '→ 11_Documentation — Dictionary & rules'),
]
for cell, text in nav:
    db[cell] = text
    db[cell].font = font_body
    db[cell].fill = fill_white
    db[cell].border = thin

db['A56'] = 'Leadership should read this sheet in under 30 seconds. Operations should open 10_Ops_Views for drill-down.'
db['A56'].font = font_small

for c in range(1, 13):
    if db.column_dimensions[get_column_letter(c)].width < 12:
        db.column_dimensions[get_column_letter(c)].width = 13
db.column_dimensions['A'].width = 22
db.column_dimensions['D'].width = 22
db.column_dimensions['H'].width = 22

# Data validations for filter cells on Dashboard Tables
dv_campus = DataValidation(type='list', formula1='" (All),SSAHE,ADYPU"', allow_blank=True)
# fix formula
dv_campus = DataValidation(type='list', formula1='"(All),SSAHE,ADYPU"', allow_blank=True)
dv_yn = DataValidation(type='list', formula1='"(All),Yes,No"', allow_blank=True)
dv_loan = DataValidation(type='list', formula1='"(All),Yes,No,Refund,DNP,Not sure"', allow_blank=True)
dt.add_data_validation(dv_campus)
dt.add_data_validation(dv_yn)
dt.add_data_validation(dv_loan)
dv_campus.add(dt['B4'])
dv_yn.add(dt['B7'])
dv_yn.add(dt['B9'])
dv_loan.add(dt['B8'])

print('Dashboard done.')

wb.save(OUT)
print('Pass2 partial save OK')
print('PASS2A_DONE')
