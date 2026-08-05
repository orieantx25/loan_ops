# -*- coding: utf-8 -*-
"""Pass 3 (light): Ops Views + Documentation — Sheets-native FILTER formulas."""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = r'uGSOT_Loan_Operations_System.xlsx'

RED = 'E31C24'
BLACK = '111111'
BORDER_C = 'E5E5E5'
MUTED = '6B6B6B'
LIGHT_RED = 'FDECEC'
GREY_HEADER = '2B2B2B'
WHITE = 'FFFFFF'

thin = Border(
    left=Side(style='thin', color=BORDER_C),
    right=Side(style='thin', color=BORDER_C),
    top=Side(style='thin', color=BORDER_C),
    bottom=Side(style='thin', color=BORDER_C),
)
fill_header = PatternFill('solid', fgColor=GREY_HEADER)
fill_light_red = PatternFill('solid', fgColor=LIGHT_RED)
fill_white = PatternFill('solid', fgColor=WHITE)

font_brand = Font(name='Calibri', size=18, bold=True, color=RED)
font_title = Font(name='Calibri', size=13, bold=True, color=BLACK)
font_h = Font(name='Calibri', size=10, bold=True, color=WHITE)
font_label = Font(name='Calibri', size=9, color=MUTED)
font_body = Font(name='Calibri', size=10, color=BLACK)
font_section = Font(name='Calibri', size=11, bold=True, color=BLACK)
font_blue = Font(name='Calibri', size=9, color='1D4ED8')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def header_row(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row, start + i, h)
        cell.fill = fill_header
        cell.font = font_h
        cell.alignment = center
        cell.border = thin

print('Opening...')
wb = load_workbook(OUT)
for name in ['10_Ops_Views', '11_Documentation']:
    if name in wb.sheetnames:
        del wb[name]

# ══════════════════════════════════════════════════════════════
print('Building 10_Ops_Views...')
ops = wb.create_sheet('10_Ops_Views')
ops.sheet_properties.tabColor = '0369A1'

ops['A1'] = 'OPERATIONS ACTION CENTER'
ops['A1'].font = font_brand
ops.merge_cells('A1:H1')
ops['A2'] = (
    'Designed for Google Sheets. Each section uses a single FILTER spill formula. '
    'In Excel 365, FILTER also works. Until then, use 04_Student_360 AutoFilter on CriticalFlag / NeedFLDG / DuplicateVendorFlag / CanonicalStage.'
)
ops['A2'].font = font_label
ops.merge_cells('A2:H2')

ops['A3'] = 'Quick counts'
ops['A3'].font = font_section
ops['A4'] = 'Pipeline (Need Loan)'
ops['B4'] = "='08_Dashboard_Tables'!B14"
ops['C4'] = 'Critical'
ops['D4'] = "='08_Dashboard_Tables'!B23"
ops['E4'] = 'FLDG'
ops['F4'] = "='08_Dashboard_Tables'!B21"
ops['G4'] = 'Dup Vendors'
ops['H4'] = "='08_Dashboard_Tables'!B24"
for col in ['B', 'D', 'F', 'H']:
    ops[f'{col}4'].font = Font(name='Calibri', size=14, bold=True, color=RED)
    ops[f'{col}4'].border = thin
    ops[f'{col}4'].fill = fill_white

# A. Top Pending
ops['A6'] = 'A. TOP PENDING STUDENTS (Pipeline, not Disbursed/Rejected/Refund/Not Required)'
ops['A6'].font = font_section
ops['A6'].fill = fill_light_red
header_row(ops, 7, [
    'StudentName', 'Campus', 'PrimaryVendor', 'Stage', 'Critical',
    'FLDG', 'Vidyalakshmi', 'Vendors', 'Days', 'Ageing', 'Case', 'Status', 'ProvID', 'Mobile'
])
ops['A8'] = (
    '=IFERROR(FILTER('
    '{\'03_Helper\'!C2:C5000,\'03_Helper\'!G2:G5000,\'03_Helper\'!T2:T5000,\'03_Helper\'!X2:X5000,'
    '\'03_Helper\'!AB2:AB5000,\'03_Helper\'!AC2:AC5000,\'03_Helper\'!AD2:AD5000,\'03_Helper\'!Q2:Q5000,'
    '\'03_Helper\'!AI2:AI5000,\'03_Helper\'!AJ2:AJ5000,\'03_Helper\'!U2:U5000,\'03_Helper\'!W2:W5000,'
    '\'03_Helper\'!E2:E5000,\'03_Helper\'!D2:D5000},'
    '(\'03_Helper\'!AK2:AK5000="Yes")*'
    '(\'03_Helper\'!X2:X5000<>"Disbursed")*(\'03_Helper\'!X2:X5000<>"Rejected")*'
    '(\'03_Helper\'!X2:X5000<>"Refund")*(\'03_Helper\'!X2:X5000<>"Not Required")'
    '),"No pending students / open in Google Sheets or Excel 365")'
)
ops['A8'].font = font_blue
ops['A8'].alignment = Alignment(wrap_text=True, vertical='top')
ops.row_dimensions[8].height = 48
ops.merge_cells('A8:N8')

ops['A9'] = 'Sort tip after spill: Critical=Yes first, then Days descending. Owner column = use Pranjal comments in Student_360.'
ops['A9'].font = font_label

# B. FLDG
ops['A11'] = 'B. NEED FLDG ATTENTION'
ops['A11'].font = font_section
ops['A11'].fill = fill_light_red
header_row(ops, 12, [
    'StudentName', 'Campus', 'ProvID', 'Stage', 'Vendor', 'Critical', 'Income', 'LoanAmt', 'Case', 'Comments'
])
ops['A13'] = (
    '=IFERROR(FILTER('
    '{\'03_Helper\'!C2:C5000,\'03_Helper\'!G2:G5000,\'03_Helper\'!E2:E5000,\'03_Helper\'!X2:X5000,'
    '\'03_Helper\'!T2:T5000,\'03_Helper\'!AB2:AB5000,\'03_Helper\'!AO2:AO5000,\'03_Helper\'!AN2:AN5000,'
    '\'03_Helper\'!U2:U5000,\'03_Helper\'!AL2:AL5000},'
    '\'03_Helper\'!AC2:AC5000="Yes"),"None / open in Sheets")'
)
ops['A13'].font = font_blue
ops.merge_cells('A13:J13')
ops.row_dimensions[13].height = 36

# C. Vidyalakshmi
ops['A15'] = 'C. NEED VIDYALAKSHMI ATTENTION'
ops['A15'].font = font_section
ops['A15'].fill = fill_light_red
header_row(ops, 16, [
    'StudentName', 'Campus', 'ProvID', 'Stage', 'Vendor', 'Critical', 'Case', 'Status'
])
ops['A17'] = (
    '=IFERROR(FILTER('
    '{\'03_Helper\'!C2:C5000,\'03_Helper\'!G2:G5000,\'03_Helper\'!E2:E5000,\'03_Helper\'!X2:X5000,'
    '\'03_Helper\'!T2:T5000,\'03_Helper\'!AB2:AB5000,\'03_Helper\'!U2:U5000,\'03_Helper\'!W2:W5000},'
    '\'03_Helper\'!AD2:AD5000="Yes"),"None / open in Sheets")'
)
ops['A17'].font = font_blue
ops.merge_cells('A17:H17')
ops.row_dimensions[17].height = 36

# D. Duplicate vendors
ops['A19'] = 'D. DUPLICATE VENDOR STUDENTS'
ops['A19'].font = font_section
ops['A19'].fill = fill_light_red
header_row(ops, 20, [
    'StudentName', 'Campus', 'VendorCount', 'VendorsApplied', 'Stage', 'Critical', 'ProvID', 'Mobile'
])
ops['A21'] = (
    '=IFERROR(FILTER('
    '{\'03_Helper\'!C2:C5000,\'03_Helper\'!G2:G5000,\'03_Helper\'!Q2:Q5000,\'03_Helper\'!S2:S5000,'
    '\'03_Helper\'!X2:X5000,\'03_Helper\'!AB2:AB5000,\'03_Helper\'!E2:E5000,\'03_Helper\'!D2:D5000},'
    '\'03_Helper\'!R2:R5000="Yes"),"None / open in Sheets")'
)
ops['A21'].font = font_blue
ops.merge_cells('A21:H21')
ops.row_dimensions[21].height = 36

# E. Critical
ops['A23'] = 'E. CRITICAL CASES'
ops['A23'].font = font_section
ops['A23'].fill = fill_light_red
header_row(ops, 24, [
    'StudentName', 'Campus', 'Stage', 'Vendor', 'RiskCategory', 'FLDG', 'Vidyalakshmi', 'ProvID'
])
ops['A25'] = (
    '=IFERROR(FILTER('
    '{\'03_Helper\'!C2:C5000,\'03_Helper\'!G2:G5000,\'03_Helper\'!X2:X5000,\'03_Helper\'!T2:T5000,'
    '\'03_Helper\'!AA2:AA5000,\'03_Helper\'!AC2:AC5000,\'03_Helper\'!AD2:AD5000,\'03_Helper\'!E2:E5000},'
    '\'03_Helper\'!AB2:AB5000="Yes"),"None / open in Sheets")'
)
ops['A25'].font = font_blue
ops.merge_cells('A25:H25')
ops.row_dimensions[25].height = 36

# F. Student lookup
ops['A27'] = 'F. STUDENT LOOKUP'
ops['A27'].font = font_section
ops['A27'].fill = fill_light_red
ops['A28'] = 'Enter Mobile or Provisional ID →'
ops['B28'] = ''
ops['B28'].fill = PatternFill('solid', fgColor='FFF3CD')
ops['B28'].border = thin
ops['C28'] = '(type here)'
ops['C28'].font = font_label

ops['A29'] = 'Name'
ops['B29'] = '=IF($B$28="","" ,IFERROR(INDEX(\'03_Helper\'!C:C,MATCH(TRUE,INDEX((\'03_Helper\'!D:D&"")=($B$28&""),0),0)),IFERROR(INDEX(\'03_Helper\'!C:C,MATCH($B$28,\'03_Helper\'!E:E,0)),"Not found")))'
ops['A30'] = 'Stage'
ops['B30'] = '=IF($B$28="","",IFERROR(INDEX(\'03_Helper\'!X:X,MATCH(TRUE,INDEX((\'03_Helper\'!D:D&"")=($B$28&""),0),0)),IFERROR(INDEX(\'03_Helper\'!X:X,MATCH($B$28,\'03_Helper\'!E:E,0)),"")))'
ops['A31'] = 'Vendors'
ops['B31'] = '=IF($B$28="","",IFERROR(INDEX(\'03_Helper\'!S:S,MATCH(TRUE,INDEX((\'03_Helper\'!D:D&"")=($B$28&""),0),0)),IFERROR(INDEX(\'03_Helper\'!S:S,MATCH($B$28,\'03_Helper\'!E:E,0)),"")))'
ops['A32'] = 'Critical'
ops['B32'] = '=IF($B$28="","",IFERROR(INDEX(\'03_Helper\'!AB:AB,MATCH(TRUE,INDEX((\'03_Helper\'!D:D&"")=($B$28&""),0),0)),IFERROR(INDEX(\'03_Helper\'!AB:AB,MATCH($B$28,\'03_Helper\'!E:E,0)),"")))'
ops['A33'] = 'Status'
ops['B33'] = '=IF($B$28="","",IFERROR(INDEX(\'03_Helper\'!W:W,MATCH(TRUE,INDEX((\'03_Helper\'!D:D&"")=($B$28&""),0),0)),IFERROR(INDEX(\'03_Helper\'!W:W,MATCH($B$28,\'03_Helper\'!E:E,0)),"")))'

for r in range(29, 34):
    ops.cell(r, 1).font = font_body
    ops.cell(r, 2).border = thin
    ops.cell(r, 2).fill = fill_white

ops['A35'] = 'G. EXCEL AUTOFILTER FALLBACK'
ops['A35'].font = font_section
ops['A36'] = 'Open 04_Student_360 → enable filters on header → filter CriticalFlag=Yes / NeedFLDG=Yes / DuplicateVendorFlag=Yes / LoanRequired=Yes. Locate any student in <10 seconds via Ctrl+F on Mobile or ProvisionalID.'
ops['A36'].font = font_body
ops.merge_cells('A36:H36')

for col in range(1, 15):
    ops.column_dimensions[get_column_letter(col)].width = 14
ops.column_dimensions['A'].width = 28
ops.column_dimensions['B'].width = 22

# ══════════════════════════════════════════════════════════════
print('Building 11_Documentation...')
doc = wb.create_sheet('11_Documentation')
doc.sheet_properties.tabColor = '666666'

doc['A1'] = 'uGSOT LOAN OPERATIONS — DOCUMENTATION'
doc['A1'].font = font_brand

sections = [
    ('', ''),
    ('1. PURPOSE', ''),
    ('', 'Formula-driven Loan Operations Management System for upGrad School of Technology. Leadership: 09_Dashboard (≤30s). Operations: 10_Ops_Views + 04_Student_360 (≤10s to find a student).'),
    ('', ''),
    ('2. SHEET MAP', ''),
    ('09_Dashboard', 'Executive UI — references 08_Dashboard_Tables only.'),
    ('01_Master_Data', 'RAW import/entry ONLY. No formulas. Paste new rows under last student.'),
    ('02_Configuration', 'Vendors, stages, campus normalization, reason keywords, ageing, cycle.'),
    ('03_Helper', 'Processing engine — StudentKey, VendorCount, CanonicalStage, Risk, Ageing.'),
    ('04_Student_360', 'Student intelligence (1 row ↔ 1 Master row; identity via StudentKey).'),
    ('05_Vendor_Analytics', 'Unique vs Applications, multi-vendor distribution, Overlap heatmap.'),
    ('06_Stage_Analytics', 'Stage counts + executive funnel conversions.'),
    ('07_Risk_Analytics', 'FLDG, Vidyalakshmi, Critical, reasons, ageing.'),
    ('08_Dashboard_Tables', 'All KPI source values. Filter dropdowns live here (B4:B9).'),
    ('10_Ops_Views', 'FILTER action lists + student lookup.'),
    ('11_Documentation', 'This sheet.'),
    ('', ''),
    ('3. STUDENT IDENTITY', ''),
    ('Primary', 'Mobile → StudentKey MOB|<n>'),
    ('Fallback', 'Provisional ID → PID|<id>'),
    ('Email', 'Not in Master (optional Merit join later).'),
    ('', ''),
    ('4. MULTI-VENDOR LOGIC', ''),
    ('Unique Students', 'Students with Shared-to-Yes for vendor (Helper flag).'),
    ('Applications', 'Equals Unique under wide-flag Master model.'),
    ('Duplicate Flag', 'VendorCount > 1'),
    ('Overlap Matrix', 'Students with both vendor flags = 1'),
    ('Never', 'Sum comma-separated Loan Status tokens as student headcount.'),
    ('', ''),
    ('5. CANONICAL STAGE ORDER', ''),
    ('', 'Refund → DNP → Disbursed → Sanctioned → Rejected → Docs Pending → Processing → Vendor Assigned → Need Loan → Loan Started → Interested → Not Required → Unclassified'),
    ('', ''),
    ('6. ADDING DATA', ''),
    ('1', 'Paste into 01_Master_Data (same headers).'),
    ('2', 'Copy last 03_Helper formula row down for new rows.'),
    ('3', 'Copy last 04_Student_360 formula row down.'),
    ('4', 'Dashboard COUNTIFs cover open columns — KPIs refresh automatically.'),
    ('Sheets', 'Upload xlsx to Google Drive → Open with Google Sheets. FILTER spills activate.'),
    ('', ''),
    ('7. EDIT RULES', ''),
    ('Allowed', 'Master Data values; Configuration; Dashboard_Tables filters B4:B9'),
    ('Forbidden', 'Hardcoding Dashboard numbers; editing Helper outputs; new pivots'),
    ('', ''),
    ('8. LEGACY PARITY TARGETS (Aug 2026 audit)', ''),
    ('Need Loan Latest=Yes', '~159'),
    ('Multi-vendor ≥2', '~62'),
    ('FLDG Yes', '~21'),
    ('Vidyalakshmi Yes', '~24'),
    ('Disbursed', '~4'),
    ('', ''),
    ('9. BRAND', ''),
    ('Primary', '#E31C24'),
    ('Black', '#111111'),
    ('Background', '#F6F6F6'),
    ('Border', '#E5E5E5'),
    ('Inspiration', 'sot.upgrad.com — minimal, premium, not a copy'),
]

for i, (a, b) in enumerate(sections, 3):
    doc.cell(i, 1, a)
    doc.cell(i, 2, b)
    if a and not b:
        doc.cell(i, 1).font = font_section
        doc.cell(i, 1).fill = fill_light_red
    else:
        doc.cell(i, 1).font = font_body
        doc.cell(i, 2).font = font_body

doc.column_dimensions['A'].width = 28
doc.column_dimensions['B'].width = 110

# Reorder
order = [
    '09_Dashboard', '01_Master_Data', '02_Configuration', '03_Helper',
    '04_Student_360', '05_Vendor_Analytics', '06_Stage_Analytics',
    '07_Risk_Analytics', '08_Dashboard_Tables', '10_Ops_Views', '11_Documentation',
]
for i, name in enumerate(order):
    current = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - current)

wb.save(OUT)
print('Saved', OUT)
print('Sheets:', wb.sheetnames)
print('PASS3_DONE')
