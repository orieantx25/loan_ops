# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('uGSOT_Loan_Operations_System.xlsx')

dt = wb['08_Dashboard_Tables']
print('Before B17:', dt['A17'].value, dt['B17'].value)
dt['B17'] = '=COUNTIF(\'03_Helper\'!X:X,"Documents Pending")'
print('After B17:', dt['B17'].value)

ops = wb['10_Ops_Views']
ops['B29'] = (
    '=IF($B$28="","",'
    'IFERROR(XLOOKUP(IFERROR(VALUE($B$28),$B$28),\'03_Helper\'!D:D,\'03_Helper\'!C:C),'
    'IFERROR(XLOOKUP($B$28,\'03_Helper\'!E:E,\'03_Helper\'!C:C),"Not found")))'
)
ops['B30'] = (
    '=IF($B$28="","",'
    'IFERROR(XLOOKUP(IFERROR(VALUE($B$28),$B$28),\'03_Helper\'!D:D,\'03_Helper\'!X:X),'
    'IFERROR(XLOOKUP($B$28,\'03_Helper\'!E:E,\'03_Helper\'!X:X),"")))'
)
ops['B31'] = (
    '=IF($B$28="","",'
    'IFERROR(XLOOKUP(IFERROR(VALUE($B$28),$B$28),\'03_Helper\'!D:D,\'03_Helper\'!S:S),'
    'IFERROR(XLOOKUP($B$28,\'03_Helper\'!E:E,\'03_Helper\'!S:S),"")))'
)
ops['B32'] = (
    '=IF($B$28="","",'
    'IFERROR(XLOOKUP(IFERROR(VALUE($B$28),$B$28),\'03_Helper\'!D:D,\'03_Helper\'!AB:AB),'
    'IFERROR(XLOOKUP($B$28,\'03_Helper\'!E:E,\'03_Helper\'!AB:AB),"")))'
)
ops['B33'] = (
    '=IF($B$28="","",'
    'IFERROR(XLOOKUP(IFERROR(VALUE($B$28),$B$28),\'03_Helper\'!D:D,\'03_Helper\'!W:W),'
    'IFERROR(XLOOKUP($B$28,\'03_Helper\'!E:E,\'03_Helper\'!W:W),"")))'
)

wb.save('uGSOT_Loan_Operations_System.xlsx')
print('Sheets:', wb.sheetnames)
print('Polished OK')
